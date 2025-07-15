import sys
import os
import json
import csv
import requests
from datetime import datetime, timedelta
from hashlib import sha256
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import urllib3

# Отключаем предупреждения для небезопасных SSL-подключений
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Проверка доступности PyQt перед импортом
try:
    from PyQt5.QtCore import Qt, QTimer
    from PyQt5.QtGui import QColor, QFont, QLinearGradient, QPainter, QBrush
    from PyQt5.QtWidgets import (
        QApplication,
        QMainWindow,
        QWidget,
        QVBoxLayout,
        QHBoxLayout,
        QPushButton,
        QLabel,
        QFrame,
        QStackedWidget,
        QTableWidget,
        QTableWidgetItem,
        QMessageBox,
        QInputDialog,
        QLineEdit,
        QDialog,
        QFormLayout,
        QDialogButtonBox,
        QDateEdit,
        QComboBox,
        QFileDialog,
        QHeaderView,
        QTextEdit,
        QSplitter,
    )
    from PyQt5.QtCore import QCoreApplication
except ImportError as e:
    print("Ошибка импорта PyQt5:", e)
    input("Нажмите Enter для выхода...")
    sys.exit(1)

# Константы
USER_DATA_FILE = "users.json"
EQUIPMENT_DATA_FILE = "equipment.json"
DEPARTMENTS = ["Туапсе", "Ильский", "Анжеро-Судженск", "Кириши"]
ROLES = {"admin": "Администратор", "metrolog": "Метролог", "guest": "Гость"}
AI_API_KEY = "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"
AI_MODEL = "deepseek/deepseek-r1"

# Данные для авторизации запросов к API Гигачат
GIGACHAT_CLIENT_ID = "a43585b3-f92b-4a57-a79a-8156056bfa84"
GIGACHAT_SCOPE = "GIGACHAT_API_PERS"
GIGACHAT_AUTH_KEY = "YTQzNTg1YjMtZjkyYi00YTU3LWE3OWEtODE1NjA1NmJmYTg0OmJlNzNhM2U0LWNmNjQtNDMyYS1iNzUzLTBlMmMwYzc5ZDkyMA=="


class AeroBackground(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.color1 = QColor(30, 136, 229)
        self.color2 = QColor(66, 165, 245)
        self.color3 = QColor(100, 181, 246)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Градиентный фон
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0, self.color1)
        gradient.setColorAt(0.5, self.color2)
        gradient.setColorAt(1, self.color3)

        painter.fillRect(self.rect(), QBrush(gradient))


def process_content(content):
    return content.replace("<think>", "").replace("</think>", "")


class AIChatWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Заголовок
        title = QLabel("Гигачат - ИИ-помощник")
        title.setStyleSheet("font: bold 16pt 'Segoe UI';")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Текстовое поле для чата
        self.chat_display = QTextEdit()
        self.chat_display.setReadOnly(True)
        self.chat_display.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border-radius: 5px;
                padding: 10px;
                font: 12pt 'Segoe UI';
            }
        """)

        # Добавляем приветственное сообщение
        self.chat_display.append(
            "<b>Гигачат:</b> Здравствуйте! Я ИИ-ассистент Гигачат. Чем могу помочь?"
        )

        # Поле ввода сообщения
        self.message_input = QTextEdit()
        self.message_input.setMaximumHeight(100)
        self.message_input.setPlaceholderText("Введите ваш вопрос...")
        self.message_input.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border-radius: 5px;
                padding: 10px;
                font: 12pt 'Segoe UI';
            }
        """)

        # Кнопка отправки
        self.send_button = QPushButton("Отправить")
        self.send_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.send_button.clicked.connect(self.send_message)

        # Индикатор статуса
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #666; font: italic 10pt 'Segoe UI';")
        self.status_label.setAlignment(Qt.AlignCenter)

        # Добавление виджетов в layout
        layout.addWidget(self.chat_display)
        layout.addWidget(self.message_input)
        layout.addWidget(self.send_button)
        layout.addWidget(self.status_label)

    def send_message(self):
        message = self.message_input.toPlainText().strip()
        if not message:
            return

        # Добавляем сообщение пользователя в чат
        self.chat_display.append(f"<b>Вы:</b> {message}")
        self.message_input.clear()

        # Отключаем кнопку отправки и показываем статус
        self.send_button.setEnabled(False)
        self.status_label.setText("Отправка запроса к Гигачат...")

        # Добавляем индикатор загрузки
        self.chat_display.append("<b>Гигачат:</b> <i>Загрузка ответа...</i>")

        # Используем QTimer для запуска запроса в отдельном потоке
        QTimer.singleShot(100, lambda: self.send_to_gigachat(message))

    def send_to_gigachat(self, message):
        try:
            # Обновляем статус
            self.status_label.setText("Получение токена авторизации...")

            # Получаем токен доступа
            auth_headers = {
                "Authorization": f"Basic {GIGACHAT_AUTH_KEY}",
                "Content-Type": "application/x-www-form-urlencoded",
                "RqUID": "123e4567-e89b-12d3-a456-426655440000",
            }

            auth_data = {
                "scope": GIGACHAT_SCOPE,
            }

            # Отправляем запрос на получение токена
            auth_response = requests.post(
                "https://ngw.devices.sberbank.ru:9443/api/v2/oauth",
                headers=auth_headers,
                data=auth_data,
                verify=False,
                timeout=30,
            )

            if auth_response.status_code != 200:
                # Обрабатываем ошибку авторизации
                error_message = f"Ошибка авторизации: {auth_response.status_code}"
                if hasattr(auth_response, "text"):
                    error_message += f" - {auth_response.text}"

                # Обновляем чат и статус
                self.update_chat_with_error(error_message)
                return

            # Получаем токен из ответа
            access_token = auth_response.json().get("access_token")
            if not access_token:
                self.update_chat_with_error("Не удалось получить токен доступа")
                return

            # Обновляем статус
            self.status_label.setText("Отправка запроса к Гигачат...")

            # Отправляем запрос к API Гигачат
            chat_headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            }

            chat_data = {
                "model": "GigaChat",
                "messages": [{"role": "user", "content": message}],
                "temperature": 0.7,
                "max_tokens": 1024,
            }

            # Отправляем запрос к API
            chat_response = requests.post(
                "https://gigachat.devices.sberbank.ru/api/v1/chat/completions",
                headers=chat_headers,
                json=chat_data,
                verify=False,
                timeout=60,
            )

            if chat_response.status_code != 200:
                # Обрабатываем ошибку API
                error_message = f"Ошибка API: {chat_response.status_code}"
                if hasattr(chat_response, "text"):
                    error_message += f" - {chat_response.text}"

                # Обновляем чат и статус
                self.update_chat_with_error(error_message)
                return

            # Обрабатываем ответ
            response_data = chat_response.json()
            if "choices" not in response_data or not response_data["choices"]:
                self.update_chat_with_error("Некорректный ответ от API")
                return

            # Получаем текст ответа
            assistant_message = response_data["choices"][0]["message"]["content"]

            # Обновляем отображение чата
            self.update_chat_with_response(assistant_message)

            # Обновляем статус
            self.status_label.setText("Ответ получен")

        except requests.exceptions.SSLError as e:
            self.update_chat_with_error(f"Ошибка SSL: {str(e)}")
        except requests.exceptions.ConnectionError as e:
            self.update_chat_with_error(f"Ошибка соединения: {str(e)}")
        except requests.exceptions.Timeout as e:
            self.update_chat_with_error(f"Таймаут запроса: {str(e)}")
        except Exception as e:
            self.update_chat_with_error(f"Непредвиденная ошибка: {str(e)}")
        finally:
            # Включаем кнопку отправки
            self.send_button.setEnabled(True)

    def update_chat_with_response(self, response_text):
        # Удаляем сообщение о загрузке
        current_text = self.chat_display.toHtml()
        loading_text = "<b>Гигачат:</b> <i>Загрузка ответа...</i>"
        if loading_text in current_text:
            current_text = current_text.replace(loading_text, "")
            self.chat_display.setHtml(current_text)

        # Добавляем ответ
        self.chat_display.append(f"<b>Гигачат:</b> {response_text}")

        # Прокручиваем до конца
        self.chat_display.moveCursor(self.chat_display.textCursor().End)

    def update_chat_with_error(self, error_message):
        # Удаляем сообщение о загрузке
        current_text = self.chat_display.toHtml()
        loading_text = "<b>Гигачат:</b> <i>Загрузка ответа...</i>"
        if loading_text in current_text:
            current_text = current_text.replace(loading_text, "")
            self.chat_display.setHtml(current_text)

        # Добавляем сообщение об ошибке
        self.chat_display.append(
            f"<b>Система:</b> <span style='color: red;'>{error_message}</span>"
        )

        # Обновляем статус
        self.status_label.setText("Произошла ошибка")

        # Прокручиваем до конца
        self.chat_display.moveCursor(self.chat_display.textCursor().End)


class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Авторизация")
        self.setFixedSize(400, 300)
        self.username = ""
        self.user_role = "guest"
        self.user_email = ""
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Вход в систему")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font: bold 18pt 'Segoe UI';")
        layout.addWidget(title)

        # Поля ввода
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Введите логин")
        form_layout.addRow("Логин:", self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)
        form_layout.addRow("Пароль:", self.password_input)

        layout.addLayout(form_layout)

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        login_btn = QPushButton("Войти")
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        login_btn.clicked.connect(self.authenticate)
        btn_layout.addWidget(login_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def authenticate(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Введите логин и пароль")
            return

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username in users:
                stored_password_hash = users[username]["password"]
                input_password_hash = sha256(password.encode("utf-8")).hexdigest()

                if stored_password_hash == input_password_hash:
                    self.username = username
                    self.user_role = users[username].get("role", "guest")
                    self.user_email = users[username].get("email", "")
                    self.accept()
                else:
                    QMessageBox.warning(self, "Ошибка", "Неверный пароль")
            else:
                QMessageBox.warning(self, "Ошибка", "Пользователь не найден")

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось загрузить данные пользователей: {str(e)}"
            )

class EquipmentDialog(QDialog):
    def __init__(self, equipment=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(
            "Добавить оборудование" if equipment is None else "Редактировать оборудование"
        )
        self.setFixedSize(500, 400)
        self.equipment = equipment
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Форма для ввода данных
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # Поля формы
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Например: Манометр ДМ-1")
        form_layout.addRow("Наименование:", self.name_input)

        self.inventory_number_input = QLineEdit()
        form_layout.addRow("Инвентарный номер:", self.inventory_number_input)

        self.serial_number_input = QLineEdit()
        form_layout.addRow("Серийный номер:", self.serial_number_input)

        self.department_combo = QComboBox()
        self.department_combo.addItems(DEPARTMENTS)
        form_layout.addRow("Подразделение:", self.department_combo)

        self.verification_date_edit = QDateEdit()
        self.verification_date_edit.setCalendarPopup(True)
        self.verification_date_edit.setDate(datetime.now().date())
        form_layout.addRow("Дата поверки:", self.verification_date_edit)

        self.next_verification_date_edit = QDateEdit()
        self.next_verification_date_edit.setCalendarPopup(True)
        self.next_verification_date_edit.setDate(
            datetime.now().date() + timedelta(days=365)
        )
        form_layout.addRow("Следующая поверка:", self.next_verification_date_edit)

        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(60)
        self.notes_input.setPlaceholderText("Дополнительная информация")
        form_layout.addRow("Примечания:", self.notes_input)

        layout.addLayout(form_layout)

        # Если редактируем существующее оборудование, заполняем поля
        if self.equipment is not None:
            self.name_input.setText(self.equipment.get("name", ""))
            self.inventory_number_input.setText(
                self.equipment.get("inventory_number", "")
            )
            self.serial_number_input.setText(self.equipment.get("serial_number", ""))
            self.department_combo.setCurrentText(
                self.equipment.get("department", DEPARTMENTS[0])
            )
            self.verification_date_edit.setDate(
                datetime.strptime(
                    self.equipment["verification_date"], "%Y-%m-%d"
                ).date()
            )
            self.next_verification_date_edit.setDate(
                datetime.strptime(
                    self.equipment["next_verification_date"], "%Y-%m-%d"
                ).date()
            )
            self.notes_input.setPlainText(self.equipment.get("notes", ""))

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_btn.clicked.connect(self.validate_and_save)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def validate_and_save(self):
        name = self.name_input.text().strip()
        inventory_number = self.inventory_number_input.text().strip()
        serial_number = self.serial_number_input.text().strip()

        if not name:
            QMessageBox.warning(self, "Ошибка", "Введите наименование оборудования")
            return

        if not inventory_number:
            QMessageBox.warning(self, "Ошибка", "Введите инвентарный номер")
            return

        verification_date = self.verification_date_edit.date().toString("yyyy-MM-dd")
        next_verification_date = self.next_verification_date_edit.date().toString(
            "yyyy-MM-dd"
        )

        equipment_data = {
            "name": name,
            "inventory_number": inventory_number,
            "serial_number": serial_number,
            "department": self.department_combo.currentText(),
            "verification_date": verification_date,
            "next_verification_date": next_verification_date,
            "notes": self.notes_input.toPlainText(),
        }

        self.equipment = equipment_data
        self.accept()

class UserManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Упр-ние пользователями")
        self.setFixedSize(600, 400)
        self.setup_ui()
        self.load_users()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Таблица пользователей
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)
        self.users_table.setHorizontalHeaderLabels(
            ["Логин", "Роль", "Email", "Действия"]
        )
        self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.users_table.verticalHeader().setVisible(False)
        self.users_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.users_table)

        # Кнопки управления
        btn_layout = QHBoxLayout()

        self.add_btn = QPushButton("Добавить пользователя")
        self.add_btn.clicked.connect(self.add_user)
        btn_layout.addWidget(self.add_btn)

        self.edit_btn = QPushButton("Редактировать")
        self.edit_btn.clicked.connect(self.edit_user)
        btn_layout.addWidget(self.edit_btn)

        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.clicked.connect(self.delete_user)
        btn_layout.addWidget(self.delete_btn)

        self.close_btn = QPushButton("Закрыть")
        self.close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.close_btn)

        layout.addLayout(btn_layout)

    def load_users(self):
        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            self.users_table.setRowCount(len(users))

            for row, (username, user_data) in enumerate(users.items()):
                # Логин
                login_item = QTableWidgetItem(username)
                login_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 0, login_item)

                # Роль
                role_item = QTableWidgetItem(
                    ROLES.get(user_data.get("role", "guest"), "Гость")
                )
                role_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 1, role_item)

                # Email
                email_item = QTableWidgetItem(user_data.get("email", ""))
                email_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 2, email_item)

                # Кнопка сброса пароля
                reset_btn = QPushButton("Сбросить пароль")
                reset_btn.clicked.connect(lambda _, u=username: self.reset_password(u))
                self.users_table.setCellWidget(row, 3, reset_btn)

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось загрузить пользователей: {str(e)}"
            )

    def add_user(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить пользователя")
        dialog.setFixedSize(400, 300)

        layout = QFormLayout(dialog)

        # Поля формы
        username_input = QLineEdit()
        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.Password)
        confirm_password_input = QLineEdit()
        confirm_password_input.setEchoMode(QLineEdit.Password)

        role_combo = QComboBox()
        role_combo.addItems(["Администратор", "Метролог", "Гость"])

        email_input = QLineEdit()

        # Добавление полей в форму
        layout.addRow("Логин:", username_input)
        layout.addRow("Пароль:", password_input)
        layout.addRow("Подтверждение пароля:", confirm_password_input)
        layout.addRow("Роль:", role_combo)
        layout.addRow("Email:", email_input)

        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(
            lambda: self.validate_and_add_user(
                dialog,
                username_input.text(),
                password_input.text(),
                confirm_password_input.text(),
                role_combo.currentText(),
                email_input.text(),
            )
        )
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        dialog.exec_()

    def validate_and_add_user(
        self, dialog, username, password, confirm_password, role, email
    ):
        if not username:
            QMessageBox.warning(self, "Ошибка", "Введите логин пользователя")
            return

        if not password:
            QMessageBox.warning(self, "Ошибка", "Введите пароль")
            return

        if password != confirm_password:
            QMessageBox.warning(self, "Ошибка", "Пароли не совпадают")
            return

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username in users:
                QMessageBox.warning(
                    self, "Ошибка", "Пользователь с таким логином уже существует"
                )
                return

            # Определяем ключ роли
            role_key = next((k for k, v in ROLES.items() if v == role), "guest")

            users[username] = {
                "password": sha256(password.encode("utf-8")).hexdigest(),
                "role": role_key,
                "email": email,
            }

            with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(users, f, indent=4, ensure_ascii=False)

            QMessageBox.information(self, "Успех", "Пользователь успешно добавлен")
            self.load_users()
            dialog.accept()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось добавить пользователя: {str(e)}"
            )

    def edit_user(self):
        selected_row = self.users_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выберите пользователя для редактирования"
            )
            return

        username = self.users_table.item(selected_row, 0).text()

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username not in users:
                QMessageBox.warning(self, "Ошибка", "Выбранный пользователь не найден")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Редактировать пользователя")
            dialog.setFixedSize(400, 250)

            layout = QFormLayout(dialog)

            # Поля формы
            role_combo = QComboBox()
            role_combo.addItems(["Администратор", "Метролог", "Гость"])

            # Устанавливаем текущую роль
            current_role = ROLES.get(users[username].get("role", "guest"), "Гость")
            role_combo.setCurrentText(current_role)

            email_input = QLineEdit(users[username].get("email", ""))

            # Добавление полей в форму
            layout.addRow("Логин:", QLabel(username))
            layout.addRow("Роль:", role_combo)
            layout.addRow("Email:", email_input)

            # Кнопки
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            buttons.accepted.connect(
                lambda: self.save_user_changes(
                    dialog, username, role_combo.currentText(), email_input.text()
                )
            )
            buttons.rejected.connect(dialog.reject)
            layout.addRow(buttons)

            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось редактировать пользователя: {str(e)}"
            )

    def save_user_changes(self, dialog, username, role, email):
        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username not in users:
                QMessageBox.warning(self, "Ошибка", "Пользователь не найден")
                return

            # Определяем ключ роли
            role_key = next((k for k, v in ROLES.items() if v == role), "guest")

            users[username]["role"] = role_key
            users[username]["email"] = email

            with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(users, f, indent=4, ensure_ascii=False)

            QMessageBox.information(self, "Успех", "Изменения сохранены")
            self.load_users()
            dialog.accept()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
            )

    def delete_user(self):
        selected_row = self.users_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите пользователя для удаления")
            return

        username = self.users_table.item(selected_row, 0).text()

        if username == "admin":
            QMessageBox.warning(
                self, "Ошибка", "Нельзя удалить администратора по умолчанию"
            )
            return

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить пользователя {username}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                    users = json.load(f)

                if username in users:
                    del users[username]

                    with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                        json.dump(users, f, indent=4, ensure_ascii=False)

                    QMessageBox.information(self, "Успех", "Пользователь удален")
                    self.load_users()

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить пользователя: {str(e)}"
                )

    def reset_password(self, username):
        if username == "admin":
            QMessageBox.warning(
                self,
                "Ошибка",
                "Для сброса пароля администратора используйте функцию изменения пароля",
            )
            return

        new_password, ok = QInputDialog.getText(
            self,
            "Сброс пароля",
            f"Введите новый пароль для пользователя {username}:",
            QLineEdit.Password,
        )

        if ok and new_password:
            try:
                with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                    users = json.load(f)

                if username in users:
                    users[username]["password"] = sha256(
                        new_password.encode("utf-8")
                    ).hexdigest()

                    with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                        json.dump(users, f, indent=4, ensure_ascii=False)

                    QMessageBox.information(self, "Успех", "Пароль успешно изменен")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка", "Пользователь не найден")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось изменить пароль: {str(e)}"
                )

class EquipmentControlApp(QMainWindow):
    def __init__(self, user_role="guest", user_email="", username=""):
        super().__init__()
        self.user_role = user_role
        self.user_email = user_email
        self.username = username
        self.equipment_data = []
        self.notification_timer = QTimer(self)
        self.notification_timer.timeout.connect(self.check_verification_dates)
        self.notification_timer.start(86400000)  # Проверка раз в день (24 часа)

        try:
            self.init_ui()
            self.load_data()
            self.check_verification_dates()
        except Exception as e:
            QMessageBox.critical(
                None,
                "Ошибка инициализации",
                f"Не удалось инициализировать приложение: {str(e)}",
            )
            sys.exit(1)

    def init_ui(self):
        self.setWindowTitle("ЛНК. ООО ЮНИКС. Контроль поверок оборудования")
        self.setMinimumSize(1200, 800)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Основной layout
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Боковая панель
        self.setup_sidebar(main_layout)

        # Область контента
        self.content_area = QStackedWidget()
        main_layout.addWidget(self.content_area)

        # Страницы приложения
        self.setup_pages()

        # Применение стилей
        self.apply_styles()

    def setup_sidebar(self, main_layout):
        sidebar = QFrame()
        sidebar.setFixedWidth(220)
        sidebar.setStyleSheet("""
            QFrame {
                background-color: #2c3e50;
                border: none;
            }
            QLabel {
                color: white;
                font: bold 16pt 'Segoe UI';
            }
            QPushButton {
                background-color: #34495e;
                color: white;
                border: none;
                border-radius: 5px;
                min-height: 40px;
                font: 12pt 'Segoe UI';
                padding: 5px 15px;
            }
            QPushButton:hover {
                background-color: #4a90e2;
            }
        """)

        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(10, 20, 10, 20)
        sidebar_layout.setSpacing(15)

        # Логотип и информация о пользователе
        logo = QLabel("Поверки")
        logo.setAlignment(Qt.AlignCenter)
        logo.setFixedHeight(60)
        sidebar_layout.addWidget(logo)

        user_info = QLabel(
            f"Пользователь: {self.username}\nРоль: {ROLES.get(self.user_role, 'Гость')}"
        )
        user_info.setAlignment(Qt.AlignCenter)
        user_info.setWordWrap(True)
        sidebar_layout.addWidget(user_info)

        # Кнопки навигации
        nav_buttons = [
            ("Главная", self.show_dashboard),
            ("Оборудование", self.show_equipment),
            ("Отчеты", self.show_reports),
        ]

        for text, handler in nav_buttons:
            btn = QPushButton(text)
            btn.clicked.connect(handler)
            sidebar_layout.addWidget(btn)

        # Кнопка управления пользователями (только для администратора)
        if self.user_role == "admin":
            self.user_management_btn = QPushButton("Упр-ние админом")
            self.user_management_btn.clicked.connect(self.manage_users)
            sidebar_layout.addWidget(self.user_management_btn)

        # Кнопка ИИ-помощника
        self.ai_btn = QPushButton("ИИ-помощник")
        self.ai_btn.clicked.connect(self.show_ai_chat)
        sidebar_layout.addWidget(self.ai_btn)

        sidebar_layout.addStretch()

        # Кнопка выхода
        btn_logout = QPushButton("Выход")
        btn_logout.setStyleSheet("background-color: #e74c3c;")
        btn_logout.clicked.connect(self.close_application)
        sidebar_layout.addWidget(btn_logout)

        main_layout.addWidget(sidebar)

    def setup_pages(self):
        # Главная страница
        self.dashboard_page = self.create_dashboard_page()
        self.content_area.addWidget(self.dashboard_page)

        # Страница оборудования
        self.equipment_page = self.create_equipment_page()
        self.content_area.addWidget(self.equipment_page)

        # Страница отчетов
        self.reports_page = self.create_reports_page()
        self.content_area.addWidget(self.reports_page)

        # Страница ИИ-помощника
        self.ai_chat_page = AIChatWidget()
        self.content_area.addWidget(self.ai_chat_page)

    def create_dashboard_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Главная")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        layout.addWidget(title)

        # Статистика
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(15)

        # Карточка с количеством оборудования
        equipment_card = QFrame()
        equipment_card.setFixedHeight(120)
        equipment_card.setStyleSheet("""
            QFrame {
                background-color: #3498db;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        equipment_layout = QVBoxLayout(equipment_card)
        equipment_layout.setContentsMargins(15, 15, 15, 15)
        self.equipment_count_label = QLabel("0")
        self.equipment_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.equipment_count_label.setAlignment(Qt.AlignCenter)
        equipment_layout.addWidget(QLabel("Всего оборудования"))
        equipment_layout.addWidget(self.equipment_count_label)
        stats_layout.addWidget(equipment_card)

        # Карточка с истекающими поверками
        expiring_card = QFrame()
        expiring_card.setFixedHeight(120)
        expiring_card.setStyleSheet("""
            QFrame {
                background-color: #e74c3c;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        expiring_layout = QVBoxLayout(expiring_card)
        expiring_layout.setContentsMargins(15, 15, 15, 15)
        self.expiring_count_label = QLabel("0")
        self.expiring_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.expiring_count_label.setAlignment(Qt.AlignCenter)
        expiring_layout.addWidget(QLabel("Истекающие поверки"))
        expiring_layout.addWidget(self.expiring_count_label)
        stats_layout.addWidget(expiring_card)

        # Карточка с просроченными поверками
        overdue_card = QFrame()
        overdue_card.setFixedHeight(120)
        overdue_card.setStyleSheet("""
            QFrame {
                background-color: #f39c12;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        overdue_layout = QVBoxLayout(overdue_card)
        overdue_layout.setContentsMargins(15, 15, 15, 15)
        self.overdue_count_label = QLabel("0")
        self.overdue_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.overdue_count_label.setAlignment(Qt.AlignCenter)
        overdue_layout.addWidget(QLabel("Просроченные поверки"))
        overdue_layout.addWidget(self.overdue_count_label)
        stats_layout.addWidget(overdue_card)

        layout.addLayout(stats_layout)

        # Таблица с ближайшими поверками
        self.verification_table = QTableWidget()
        self.verification_table.setColumnCount(5)
        self.verification_table.setHorizontalHeaderLabels(
            ["Наименование", "Инв. номер", "Подразделение", "Дата поверки", "Статус"]
        )
        self.verification_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.verification_table.verticalHeader().setVisible(False)
        self.verification_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(QLabel("Ближайшие поверки:"))
        layout.addWidget(self.verification_table)

        return page

    def create_equipment_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_layout = QHBoxLayout()
        title = QLabel("Оборудование")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        header_layout.addWidget(title)

        # Кнопки управления (только для администратора и метролога)
        if self.user_role in ["admin", "metrolog"]:
            btn_layout = QHBoxLayout()
            btn_layout.setSpacing(10)

            add_btn = QPushButton("Добавить")
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            add_btn.clicked.connect(self.add_equipment)
            btn_layout.addWidget(add_btn)

            edit_btn = QPushButton("Редактировать")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #0b7dda;
                }
            """)
            edit_btn.clicked.connect(self.edit_equipment)
            btn_layout.addWidget(edit_btn)

            delete_btn = QPushButton("Удалить")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #d32f2f;
                }
            """)
            delete_btn.clicked.connect(self.delete_equipment)
            btn_layout.addWidget(delete_btn)

            header_layout.addLayout(btn_layout)

        layout.addLayout(header_layout)

        # Фильтры
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.department_filter = QComboBox()
        self.department_filter.addItem("Все подразделения")
        for dept in DEPARTMENTS:
            self.department_filter.addItem(dept)
        self.department_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Подразделение:"))
        filter_layout.addWidget(self.department_filter)

        self.status_filter = QComboBox()
        self.status_filter.addItem("Все статусы", "")
        self.status_filter.addItem("Активные", "active")
        self.status_filter.addItem("Истекающие", "expiring")
        self.status_filter.addItem("Просроченные", "overdue")
        self.status_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Статус:"))
        filter_layout.addWidget(self.status_filter)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по наименованию или номеру...")
        self.search_input.textChanged.connect(self.apply_filters)
        search_layout.addWidget(self.search_input)

        export_btn = QPushButton("Экспорт")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        export_btn.clicked.connect(self.export_equipment)
        search_layout.addWidget(export_btn)

        filter_layout.addLayout(search_layout)
        layout.addLayout(filter_layout)

        # Таблица оборудования
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(7)
        self.equipment_table.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Инв. номер",
                "Сер. номер",
                "Подразделение",
                "Дата поверки",
                "След. поверка",
                "Статус",
            ]
        )
        self.equipment_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.equipment_table.verticalHeader().setVisible(False)
        self.equipment_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.equipment_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.equipment_table.setSelectionMode(QTableWidget.SingleSelection)
        self.equipment_table.setSortingEnabled(True)
        layout.addWidget(self.equipment_table)

        return page

    def create_reports_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Отчеты")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        layout.addWidget(title)

        # Кнопки отчетов
        reports_layout = QHBoxLayout()
        reports_layout.setSpacing(15)

        # Кнопка для отчета по истекающим поверкам
        expiring_btn = QPushButton("Истекающие поверки")
        expiring_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        expiring_btn.clicked.connect(lambda: self.generate_report("expiring"))
        reports_layout.addWidget(expiring_btn)

        # Кнопка для отчета по просроченным поверкам
        overdue_btn = QPushButton("Просроченные поверки")
        overdue_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        overdue_btn.clicked.connect(lambda: self.generate_report("overdue"))
        reports_layout.addWidget(overdue_btn)

        # Кнопка для полного отчета
        full_report_btn = QPushButton("Полный отчет")
        full_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        full_report_btn.clicked.connect(lambda: self.generate_report("full"))
        reports_layout.addWidget(full_report_btn)

        layout.addLayout(reports_layout)

        # Параметры отчета
        params_layout = QFormLayout()
        params_layout.setSpacing(15)

        # Выбор подразделения
        self.report_department_combo = QComboBox()
        self.report_department_combo.addItem("Все подразделения")
        self.report_department_combo.addItems(DEPARTMENTS)
        params_layout.addRow("Подразделение:", self.report_department_combo)

        # Выбор формата экспорта
        self.export_format_combo = QComboBox()
        self.export_format_combo.addItem("Excel", "xlsx")
        self.export_format_combo.addItem("CSV", "csv")
        self.export_format_combo.addItem("PDF", "pdf")
        params_layout.addRow("Формат:", self.export_format_combo)

        layout.addLayout(params_layout)

        # Область предпросмотра отчета
        self.report_preview = QTableWidget()
        self.report_preview.setColumnCount(6)
        self.report_preview.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Инв. номер",
                "Подразделение",
                "Дата поверки",
                "След. поверка",
                "Статус",
            ]
        )
        self.report_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.report_preview.verticalHeader().setVisible(False)
        self.report_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(QLabel("Предпросмотр:"))
        layout.addWidget(self.report_preview)

        return page

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                font: 11pt 'Segoe UI';
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 5px;
                font: bold 11pt 'Segoe UI';
            }
            QLineEdit, QComboBox {
                padding: 5px;
                font: 11pt 'Segoe UI';
                border: 1px solid #ddd;
                border-radius: 3px;
            }
            QLabel {
                font: 11pt 'Segoe UI';
            }
        """)

    def load_data(self):
        try:
            if not os.path.exists(EQUIPMENT_DATA_FILE):
                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump([], f)

            with open(EQUIPMENT_DATA_FILE, "r", encoding="utf-8") as f:
                self.equipment_data = json.load(f)

            self.update_equipment_table()
            self.update_dashboard()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка загрузки",
                f"Не удалось загрузить данные оборудования: {str(e)}",
            )

    def update_equipment_table(self):
        self.equipment_table.setRowCount(len(self.equipment_data))

        for row, equipment in enumerate(self.equipment_data):
            # Наименование
            name_item = QTableWidgetItem(equipment["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.equipment_table.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(equipment["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 1, inv_item)

            # Серийный номер
            serial_item = QTableWidgetItem(equipment["serial_number"])
            serial_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 2, serial_item)

            # Подразделение
            dept_item = QTableWidgetItem(equipment["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 3, dept_item)

            # Дата поверки
            verif_date = datetime.strptime(
                equipment["verification_date"], "%Y-%m-%d"
            ).date()
            verif_item = QTableWidgetItem(verif_date.strftime("%d.%m.%Y"))
            verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 4, verif_item)

            # Следующая поверка
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            next_verif_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            next_verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 5, next_verif_item)

            # Статус
            status, color = self.get_verification_status(next_verif_date)
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(color))
            self.equipment_table.setItem(row, 6, status_item)

    def update_dashboard(self):
        # Общее количество оборудования
        self.equipment_count_label.setText(str(len(self.equipment_data)))

        # Количество истекающих и просроченных поверок
        expiring_count = 0
        overdue_count = 0
        today = datetime.now().date()

        verification_data = []

        for equipment in self.equipment_data:
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if days_left < 0:
                overdue_count += 1
                status = "Просрочено"
                color = "#e74c3c"
            elif days_left <= 30:
                expiring_count += 1
                status = f"Истекает через {days_left} дней"
                color = "#f39c12"
            else:
                status = f"Активно ({days_left} дней)"
                color = "#2ecc71"

            verification_data.append(
                {
                    "equipment": equipment,
                    "days_left": days_left,
                    "status": status,
                    "color": color,
                }
            )

        self.expiring_count_label.setText(str(expiring_count))
        self.overdue_count_label.setText(str(overdue_count))

        # Сортируем по дням до поверки (ближайшие сверху)
        verification_data.sort(key=lambda x: x["days_left"])

        # Обновляем таблицу ближайших поверок
        self.verification_table.setRowCount(min(10, len(verification_data)))

        for row, data in enumerate(verification_data[:10]):
            eq = data["equipment"]

            # Наименование
            name_item = QTableWidgetItem(eq["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.verification_table.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(eq["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 1, inv_item)

            # Подразделение
            dept_item = QTableWidgetItem(eq["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 2, dept_item)

            # Дата следующей поверки
            next_verif_date = datetime.strptime(
                eq["next_verification_date"], "%Y-%m-%d"
            ).date()
            date_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            date_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 3, date_item)

            # Статус
            status_item = QTableWidgetItem(data["status"])
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(data["color"]))
            self.verification_table.setItem(row, 4, status_item)

    def get_verification_status(self, next_verif_date):
        today = datetime.now().date()
        days_left = (next_verif_date - today).days

        if days_left < 0:
            return "Просрочено", "#e74c3c"
        elif days_left <= 30:
            return f"Истекает ({days_left} дн.)", "#f39c12"
        else:
            return f"Активно ({days_left} дн.)", "#2ecc71"

    def apply_filters(self):
        department_filter = self.department_filter.currentText()
        status_filter = self.status_filter.currentData()
        search_text = self.search_input.text().lower()
        today = datetime.now().date()

        for row in range(self.equipment_table.rowCount()):
            should_show = True

            # Фильтр по подразделению
            if department_filter != "Все подразделения":
                dept_item = self.equipment_table.item(row, 3)
                if dept_item.text() != department_filter:
                    should_show = False

            # Фильтр по статусу
            if status_filter and should_show:
                next_verif_item = self.equipment_table.item(row, 5)
                next_verif_date = datetime.strptime(
                    next_verif_item.text(), "%d.%m.%Y"
                ).date()
                days_left = (next_verif_date - today).days

                if status_filter == "active" and days_left <= 30:
                    should_show = False
                elif status_filter == "expiring" and not (0 <= days_left <= 30):
                    should_show = False
                elif status_filter == "overdue" and days_left >= 0:
                    should_show = False

            # Фильтр по поиску
            if should_show and search_text:
                name_item = self.equipment_table.item(row, 0)
                inv_item = self.equipment_table.item(row, 1)

                if (
                    search_text not in name_item.text().lower()
                    and search_text not in inv_item.text().lower()
                ):
                    should_show = False

            self.equipment_table.setRowHidden(row, not should_show)

    def add_equipment(self):
        dialog = EquipmentDialog()
        if dialog.exec_() == QDialog.Accepted:
            try:
                self.equipment_data.append(dialog.equipment)

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Оборудование успешно добавлено")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось добавить оборудование: {str(e)}"
                )

    def edit_equipment(self):
        selected_row = self.equipment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выберите оборудование для редактирования"
            )
            return

        # Получаем инвентарный номер и серийный номер для идентификации оборудования
        inv_number = self.equipment_table.item(selected_row, 1).text()
        serial_number = self.equipment_table.item(selected_row, 2).text()

        # Ищем соответствующее оборудование в массиве данных
        found_index = -1
        for i, eq in enumerate(self.equipment_data):
            if (
                eq["inventory_number"] == inv_number
                and eq["serial_number"] == serial_number
            ):
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранное оборудование не найдено в базе данных"
            )
            return

        equipment = self.equipment_data[found_index]
        dialog = EquipmentDialog(equipment, self)

        if dialog.exec_() == QDialog.Accepted:
            try:
                self.equipment_data[found_index] = dialog.equipment

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Изменения сохранены")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
                )

    def delete_equipment(self):
        selected_row = self.equipment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите оборудование для удаления")
            return

        # Получаем инвентарный номер и серийный номер для идентификации оборудования
        inv_number = self.equipment_table.item(selected_row, 1).text()
        serial_number = self.equipment_table.item(selected_row, 2).text()

        # Ищем соответствующее оборудование в массиве данных
        found_index = -1
        for i, eq in enumerate(self.equipment_data):
            if (
                eq["inventory_number"] == inv_number
                and eq["serial_number"] == serial_number
            ):
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранное оборудование не найдено в базе данных"
            )
            return

        equipment = self.equipment_data[found_index]

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить оборудование {equipment['name']} (инв. № {equipment['inventory_number']})?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                del self.equipment_data[found_index]

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Оборудование удалено")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить оборудование: {str(e)}"
                )

    def export_equipment(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт данных",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
            options=options,
        )

        if not file_name:
            return

        try:
            if file_name.endswith(".xlsx"):
                self.export_to_excel(file_name)
            else:
                self.export_to_csv(file_name)

            QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось экспортировать данные: {str(e)}"
            )

    def export_to_excel(self, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Оборудование"

        # Заголовки
        headers = [
            "Наименование",
            "Инвентарный номер",
            "Серийный номер",
            "Подразделение",
            "Дата поверки",
            "Следующая поверка",
            "Статус",
            "Примечания",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Данные
        today = datetime.now().date()

        for row, equipment in enumerate(self.equipment_data, 2):
            sheet.cell(row=row, column=1, value=equipment["name"])

            sheet.cell(row=row, column=2, value=equipment["inventory_number"])
            sheet.cell(row=row, column=3, value=equipment["serial_number"])
            sheet.cell(row=row, column=4, value=equipment["department"])

            verif_date = datetime.strptime(
                equipment["verification_date"], "%Y-%m-%d"
            ).date()
            sheet.cell(row=row, column=5, value=verif_date.strftime("%d.%m.%Y"))

            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            sheet.cell(row=row, column=6, value=next_verif_date.strftime("%d.%m.%Y"))

            days_left = (next_verif_date - today).days
            if days_left < 0:
                status = "Просрочено"
            elif days_left <= 30:
                status = f"Истекает через {days_left} дней"
            else:
                status = f"Активно ({days_left} дней)"

            sheet.cell(row=row, column=7, value=status)
            sheet.cell(row=row, column=8, value=equipment.get("notes", ""))

        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Границы для всех ячеек
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        workbook.save(file_path)

    def export_to_csv(self, file_path):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            # Заголовки
            writer.writerow(
                [
                    "Наименование",
                    "Инвентарный номер",
                    "Серийный номер",
                    "Подразделение",
                    "Дата поверки",
                    "Следующая поверка",
                    "Статус",
                    "Примечания",
                ]
            )

            # Данные
            today = datetime.now().date()

            for equipment in self.equipment_data:
                next_verif_date = datetime.strptime(
                    equipment["next_verification_date"], "%Y-%m-%d"
                ).date()
                days_left = (next_verif_date - today).days

                if days_left < 0:
                    status = "Просрочено"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                else:
                    status = f"Активно ({days_left} дней)"

                writer.writerow(
                    [
                        equipment["name"],
                        equipment["inventory_number"],
                        equipment["serial_number"],
                        equipment["department"],
                        datetime.strptime(
                            equipment["verification_date"], "%Y-%m-%d"
                        ).strftime("%d.%m.%Y"),
                        next_verif_date.strftime("%d.%m.%Y"),
                        status,
                        equipment.get("notes", ""),
                    ]
                )

    def generate_report(self, report_type):
        today = datetime.now().date()
        filtered_data = []

        department_filter = self.report_department_combo.currentText()
        if department_filter == "Все подразделения":
            department_filter = None

        for equipment in self.equipment_data:
            if department_filter and equipment["department"] != department_filter:
                continue

            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if report_type == "expiring" and not (0 <= days_left <= 30):
                continue
            elif report_type == "overdue" and days_left >= 0:
                continue

            filtered_data.append({"equipment": equipment, "days_left": days_left})

        # Сортируем данные
        if report_type in ["expiring", "overdue"]:
            filtered_data.sort(key=lambda x: x["days_left"])
        else:
            filtered_data.sort(key=lambda x: x["equipment"]["department"])

        # Обновляем предпросмотр
        self.report_preview.setRowCount(len(filtered_data))

        for row, data in enumerate(filtered_data):
            eq = data["equipment"]

            # Наименование
            name_item = QTableWidgetItem(eq["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.report_preview.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(eq["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.report_preview.setItem(row, 1, inv_item)

            # Подразделение
            dept_item = QTableWidgetItem(eq["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.report_preview.setItem(row, 2, dept_item)

            # Дата поверки
            verif_date = datetime.strptime(eq["verification_date"], "%Y-%m-%d").date()
            verif_item = QTableWidgetItem(verif_date.strftime("%d.%m.%Y"))
            verif_item.setTextAlignment(Qt.AlignCenter)
            self.report_preview.setItem(row, 3, verif_item)

            # Следующая поверка
            next_verif_date = datetime.strptime(
                eq["next_verification_date"], "%Y-%m-%d"
            ).date()
            next_verif_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            next_verif_item.setTextAlignment(Qt.AlignCenter)
            self.report_preview.setItem(row, 4, next_verif_item)

            # Статус
            if data["days_left"] < 0:
                status = "Просрочено"
                color = "#e74c3c"
            elif data["days_left"] <= 30:
                status = f"Истекает через {data['days_left']} дней"
                color = "#f39c12"
            else:
                status = f"Активно ({data['days_left']} дней)"
                color = "#2ecc71"

            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(color))
            self.report_preview.setItem(row, 5, status_item)

    def check_verification_dates(self):
        today = datetime.now().date()
        expiring_soon = []
        overdue = []

        for equipment in self.equipment_data:
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if days_left < 0:
                overdue.append(equipment)
            elif days_left <= 7:  # Уведомляем за неделю до поверки
                expiring_soon.append(equipment)

        if overdue or expiring_soon:
            message = ""

            if overdue:
                message += "<b>Просроченные поверки:</b><br>"
                for eq in overdue:
                    message += f"- {eq['name']} (инв. № {eq['inventory_number']}, {eq['department']})<br>"
                message += "<br>"

            if expiring_soon:
                message += "<b>Скоро истекают поверки:</b><br>"
                for eq in expiring_soon:
                    next_date = datetime.strptime(
                        eq["next_verification_date"], "%Y-%m-%d"
                    ).date()
                    days_left = (next_date - today).days
                    message += f"- {eq['name']} (инв. № {eq['inventory_number']}, {eq['department']}) - осталось {days_left} дней<br>"

            # Показываем уведомление только если пользователь активен
            if self.isActiveWindow():
                QMessageBox.warning(self, "Проверка поверок", message, QMessageBox.Ok)

            # Отправляем уведомление на email, если он указан
            if self.user_email and (overdue or expiring_soon):
                self.send_email_notification(overdue, expiring_soon)

    def send_email_notification(self, overdue, expiring_soon):
        # В реальном приложении здесь должна быть реализация отправки email
        # Это примерная заглушка для демонстрации
        print(f"Уведомление отправлено на email: {self.user_email}")
        print("Просроченные поверки:", overdue)
        print("Скоро истекают:", expiring_soon)

    def show_dashboard(self):
        self.content_area.setCurrentWidget(self.dashboard_page)

    def show_equipment(self):
        self.content_area.setCurrentWidget(self.equipment_page)

    def show_reports(self):
        self.content_area.setCurrentWidget(self.reports_page)

    def show_ai_chat(self):
        self.content_area.setCurrentWidget(self.ai_chat_page)

    def manage_users(self):
        if self.user_role != "admin":
            QMessageBox.warning(
                self, "Ошибка", "Недостаточно прав для управления пользователями"
            )
            return

        dialog = UserManagementDialog(self)
        dialog.exec_()

    def close_application(self):
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите выйти?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            QCoreApplication.quit()

def initialize_data_files():
    # Создаем файл пользователей, если его нет
    if not os.path.exists(USER_DATA_FILE):
        default_users = {
            "admin": {
                "password": sha256("admin123".encode("utf-8")).hexdigest(),
                "role": "admin",
                "email": "admin@example.com",
            }
        }

        with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(default_users, f, indent=4, ensure_ascii=False)

    # Создаем файл оборудования, если его нет
    if not os.path.exists(EQUIPMENT_DATA_FILE):
        with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)

def main():
    # Инициализация файлов данных
    initialize_data_files()

    # Создание приложения
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Авторизация пользователя
    login_dialog = LoginDialog()
    if login_dialog.exec_() == QDialog.Accepted:
        # Запуск основного приложения
        main_app = EquipmentControlApp(
            user_role=login_dialog.user_role,
            user_email=login_dialog.user_email,
            username=login_dialog.username,
        )
        main_app.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)

if __name__ == "__main__":
    main()