import sys
import os
import json
import csv
import requests
import shutil
from datetime import datetime, timedelta
from hashlib import sha256
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import urllib3

# Отключаем предупреждения для небезопасных SSL-подключений
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Проверка доступности PyQt перед импортом
try:
    from PyQt5.QtCore import Qt, QTimer, QUrl
    from PyQt5.QtGui import QColor, QFont, QLinearGradient, QPainter, QBrush, QDesktopServices
    from PyQt5.QtWidgets import (
        QApplication,
        QMainWindow,
        QWidget,
        QVBoxLayout,
        QHBoxLayout,
        QPushButton,
        QLabel,
        QFrame,
        QStackedWidget,
        QTableWidget,
        QTableWidgetItem,
        QMessageBox,
        QInputDialog,
        QLineEdit,
        QDialog,
        QFormLayout,
        QDialogButtonBox,
        QDateEdit,
        QComboBox,
        QFileDialog,
        QHeaderView,
        QTextEdit,
        QSplitter,
        QScrollArea,
        QGridLayout,
        QListWidget,
        QListWidgetItem,
        QTabWidget,
    )
    from PyQt5.QtCore import QCoreApplication
    from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
except ImportError as e:
    print("Ошибка импорта PyQt5:", e)
    input("Нажмите Enter для выхода...")
    sys.exit(1)

# Константы
USER_DATA_FILE = "users.json"
EQUIPMENT_DATA_FILE = "equipment.json"
SPECIALISTS_DATA_FILE = "specialists.json"
OBJECTS_DATA_FILE = "objects.json"
CONFIG_FILE = "config.json"
DOCS_FOLDER = "equipment_docs"
DEPARTMENTS = ["Туапсе", "Ильский", "Анжеро-Судженск", "Кириши"]
ROLES = {
    "admin": "Администратор", 
    "metrolog": "Метролог", 
    "expert": "Эксперт",
    "lab": "Лаборант",
    "manager": "Руководитель",
    "guest": "Гость"
}
CONTROL_METHODS = [
    "УЗК", "РК", "ВИК", "ПВК", "МК", "ЭК", "АЭ", "ТК", "ВК", "ЭМК"
]
AI_API_KEY = "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"
AI_MODEL = "deepseek/deepseek-r1"

# Данные для авторизации запросов к API Гигачат
GIGACHAT_CLIENT_ID = "a43585b3-f92b-4a57-a79a-8156056bfa84"
GIGACHAT_SCOPE = "GIGACHAT_API_PERS"
GIGACHAT_AUTH_KEY = "YTQzNTg1YjMtZjkyYi00YTU3LWE3OWEtODE1NjA1NmJmYTg0OmJlNzNhM2U0LWNmNjQtNDMyYS1iNzUzLTBlMmMwYzc5ZDkyMA=="


class AeroBackground(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.color1 = QColor(30, 136, 229)
        self.color2 = QColor(66, 165, 245)
        self.color3 = QColor(100, 181, 246)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Градиентный фон
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0, self.color1)
        gradient.setColorAt(0.5, self.color2)
        gradient.setColorAt(1, self.color3)

        painter.fillRect(self.rect(), QBrush(gradient))


def process_content(content):
    return content.replace("<think>", "").replace("</think>", "")


class AIChatWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Заголовок
        title = QLabel("Гигачат - ИИ-помощник")
        title.setStyleSheet("font: bold 16pt 'Segoe UI';")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Текстовое поле для чата
        self.chat_display = QTextEdit()
        self.chat_display.setReadOnly(True)
        self.chat_display.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border-radius: 5px;
                padding: 10px;
                font: 12pt 'Segoe UI';
            }
        """)

        # Добавляем приветственное сообщение
        self.chat_display.append(
            "<b>Гигачат:</b> Здравствуйте! Я ИИ-ассистент Гигачат. Чем могу помочь?"
        )

        # Поле ввода сообщения
        self.message_input = QTextEdit()
        self.message_input.setMaximumHeight(100)
        self.message_input.setPlaceholderText("Введите ваш вопрос...")
        self.message_input.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border-radius: 5px;
                padding: 10px;
                font: 12pt 'Segoe UI';
            }
        """)

        # Кнопка отправки
        self.send_button = QPushButton("Отправить")
        self.send_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.send_button.clicked.connect(self.send_message)

        # Индикатор статуса
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #666; font: italic 10pt 'Segoe UI';")
        self.status_label.setAlignment(Qt.AlignCenter)

        # Добавление виджетов в layout
        layout.addWidget(self.chat_display)
        layout.addWidget(self.message_input)
        layout.addWidget(self.send_button)
        layout.addWidget(self.status_label)

    def send_message(self):
        message = self.message_input.toPlainText().strip()
        if not message:
            return

        # Добавляем сообщение пользователя в чат
        self.chat_display.append(f"<b>Вы:</b> {message}")
        self.message_input.clear()

        # Отключаем кнопку отправки и показываем статус
        self.send_button.setEnabled(False)
        self.status_label.setText("Отправка запроса к Гигачат...")

        # Добавляем индикатор загрузки
        self.chat_display.append("<b>Гигачат:</b> <i>Загрузка ответа...</i>")

        # Используем QTimer для запуска запроса в отдельном потоке
        QTimer.singleShot(100, lambda: self.send_to_gigachat(message))

    def send_to_gigachat(self, message):
        try:
            # Обновляем статус
            self.status_label.setText("Получение токена авторизации...")

            # Получаем токен доступа
            auth_headers = {
                "Authorization": f"Basic {GIGACHAT_AUTH_KEY}",
                "Content-Type": "application/x-www-form-urlencoded",
                "RqUID": "123e4567-e89b-12d3-a456-426655440000",
            }

            auth_data = {
                "scope": GIGACHAT_SCOPE,
            }

            # Отправляем запрос на получение токена
            auth_response = requests.post(
                "https://ngw.devices.sberbank.ru:9443/api/v2/oauth",
                headers=auth_headers,
                data=auth_data,
                verify=False,
                timeout=30,
            )

            if auth_response.status_code != 200:
                # Обрабатываем ошибку авторизации
                error_message = f"Ошибка авторизации: {auth_response.status_code}"
                if hasattr(auth_response, "text"):
                    error_message += f" - {auth_response.text}"

                # Обновляем чат и статус
                self.update_chat_with_error(error_message)
                return

            # Получаем токен из ответа
            access_token = auth_response.json().get("access_token")
            if not access_token:
                self.update_chat_with_error("Не удалось получить токен доступа")
                return

            # Обновляем статус
            self.status_label.setText("Отправка запроса к Гигачат...")

            # Отправляем запрос к API Гигачат
            chat_headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            }

            chat_data = {
                "model": "GigaChat",
                "messages": [{"role": "user", "content": message}],
                "temperature": 0.7,
                "max_tokens": 1024,
            }

            # Отправляем запрос к API
            chat_response = requests.post(
                "https://gigachat.devices.sberbank.ru/api/v1/chat/completions",
                headers=chat_headers,
                json=chat_data,
                verify=False,
                timeout=60,
            )

            if chat_response.status_code != 200:
                # Обрабатываем ошибку API
                error_message = f"Ошибка API: {chat_response.status_code}"
                if hasattr(chat_response, "text"):
                    error_message += f" - {chat_response.text}"

                # Обновляем чат и статус
                self.update_chat_with_error(error_message)
                return

            # Обрабатываем ответ
            response_data = chat_response.json()
            if "choices" not in response_data or not response_data["choices"]:
                self.update_chat_with_error("Некорректный ответ от API")
                return

            # Получаем текст ответа
            assistant_message = response_data["choices"][0]["message"]["content"]

            # Обновляем отображение чата
            self.update_chat_with_response(assistant_message)

            # Обновляем статус
            self.status_label.setText("Ответ получен")

        except requests.exceptions.SSLError as e:
            self.update_chat_with_error(f"Ошибка SSL: {str(e)}")
        except requests.exceptions.ConnectionError as e:
            self.update_chat_with_error(f"Ошибка соединения: {str(e)}")
        except requests.exceptions.Timeout as e:
            self.update_chat_with_error(f"Таймаут запроса: {str(e)}")
        except Exception as e:
            self.update_chat_with_error(f"Непредвиденная ошибка: {str(e)}")
        finally:
            # Включаем кнопку отправки
            self.send_button.setEnabled(True)

    def update_chat_with_response(self, response_text):
        # Удаляем сообщение о загрузке
        current_text = self.chat_display.toHtml()
        loading_text = "<b>Гигачат:</b> <i>Загрузка ответа...</i>"
        if loading_text in current_text:
            current_text = current_text.replace(loading_text, "")
            self.chat_display.setHtml(current_text)

        # Добавляем ответ
        self.chat_display.append(f"<b>Гигачат:</b> {response_text}")

        # Прокручиваем до конца
        self.chat_display.moveCursor(self.chat_display.textCursor().End)

    def update_chat_with_error(self, error_message):
        # Удаляем сообщение о загрузке
        current_text = self.chat_display.toHtml()
        loading_text = "<b>Гигачат:</b> <i>Загрузка ответа...</i>"
        if loading_text in current_text:
            current_text = current_text.replace(loading_text, "")
            self.chat_display.setHtml(current_text)

        # Добавляем сообщение об ошибке
        self.chat_display.append(
            f"<b>Система:</b> <span style='color: red;'>{error_message}</span>"
        )

        # Обновляем статус
        self.status_label.setText("Произошла ошибка")

        # Прокручиваем до конца
        self.chat_display.moveCursor(self.chat_display.textCursor().End)


class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Авторизация")
        self.setFixedSize(400, 300)
        self.username = ""
        self.user_role = "guest"
        self.user_email = ""
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Вход в систему")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font: bold 18pt 'Segoe UI';")
        layout.addWidget(title)

        # Поля ввода
        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Введите логин")
        form_layout.addRow("Логин:", self.username_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль")
        self.password_input.setEchoMode(QLineEdit.Password)
        form_layout.addRow("Пароль:", self.password_input)

        layout.addLayout(form_layout)

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        login_btn = QPushButton("Войти")
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        login_btn.clicked.connect(self.authenticate)
        btn_layout.addWidget(login_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def authenticate(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(self, "Ошибка", "Введите логин и пароль")
            return

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username in users:
                stored_password_hash = users[username]["password"]
                input_password_hash = sha256(password.encode("utf-8")).hexdigest()

                if stored_password_hash == input_password_hash:
                    self.username = username
                    self.user_role = users[username].get("role", "guest")
                    self.user_email = users[username].get("email", "")
                    self.accept()
                else:
                    QMessageBox.warning(self, "Ошибка", "Неверный пароль")
            else:
                QMessageBox.warning(self, "Ошибка", "Пользователь не найден")

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось загрузить данные пользователей: {str(e)}"
            )

class EquipmentDialog(QDialog):
    def __init__(self, equipment=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(
            "Добавить оборудование" if equipment is None else "Редактировать оборудование"
        )
        self.setFixedSize(600, 500)
        self.equipment = equipment or {"documents": []}
        if "documents" not in self.equipment:
            self.equipment["documents"] = []
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Форма для ввода данных
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # Поля формы
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Например: Манометр ДМ-1")
        form_layout.addRow("Наименование:", self.name_input)

        self.inventory_number_input = QLineEdit()
        form_layout.addRow("Инвентарный номер:", self.inventory_number_input)

        self.serial_number_input = QLineEdit()
        form_layout.addRow("Серийный номер:", self.serial_number_input)

        self.department_combo = QComboBox()
        self.department_combo.addItems(DEPARTMENTS)
        form_layout.addRow("Подразделение:", self.department_combo)

        self.verification_date_edit = QDateEdit()
        self.verification_date_edit.setCalendarPopup(True)
        self.verification_date_edit.setDate(datetime.now().date())
        form_layout.addRow("Дата поверки:", self.verification_date_edit)

        self.next_verification_date_edit = QDateEdit()
        self.next_verification_date_edit.setCalendarPopup(True)
        self.next_verification_date_edit.setDate(
            datetime.now().date() + timedelta(days=365)
        )
        form_layout.addRow("Следующая поверка:", self.next_verification_date_edit)

        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(60)
        self.notes_input.setPlaceholderText("Дополнительная информация")
        form_layout.addRow("Примечания:", self.notes_input)

        layout.addLayout(form_layout)

        # Список прикрепленных документов
        docs_group = QFrame()
        docs_layout = QVBoxLayout(docs_group)

        docs_label = QLabel("Прикрепленные документы:")
        docs_label.setStyleSheet("font-weight: bold;")
        docs_layout.addWidget(docs_label)

        self.docs_list = QListWidget()
        self.docs_list.setMaximumHeight(100)
        docs_layout.addWidget(self.docs_list)

        docs_buttons = QHBoxLayout()
        
        self.attach_doc_button = QPushButton("Прикрепить файл")
        self.attach_doc_button.clicked.connect(self.attach_document)
        docs_buttons.addWidget(self.attach_doc_button)
        
        self.remove_doc_button = QPushButton("Удалить")
        self.remove_doc_button.clicked.connect(self.remove_document)
        docs_buttons.addWidget(self.remove_doc_button)
        
        docs_layout.addLayout(docs_buttons)
        layout.addWidget(docs_group)

        # Если редактируем существующее оборудование, заполняем поля
        if self.equipment is not None:
            self.name_input.setText(self.equipment.get("name", ""))
            self.inventory_number_input.setText(
                self.equipment.get("inventory_number", "")
            )
            self.serial_number_input.setText(self.equipment.get("serial_number", ""))
            self.department_combo.setCurrentText(
                self.equipment.get("department", DEPARTMENTS[0])
            )
            
            if "verification_date" in self.equipment:
                self.verification_date_edit.setDate(
                    datetime.strptime(
                        self.equipment["verification_date"], "%Y-%m-%d"
                    ).date()
                )
                
            if "next_verification_date" in self.equipment:
                self.next_verification_date_edit.setDate(
                    datetime.strptime(
                        self.equipment["next_verification_date"], "%Y-%m-%d"
                    ).date()
                )
                
            self.notes_input.setPlainText(self.equipment.get("notes", ""))
            
            # Загружаем список документов
            for doc in self.equipment.get("documents", []):
                self.docs_list.addItem(doc["name"])

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_btn.clicked.connect(self.validate_and_save)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def attach_document(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Выберите документ", 
            "", 
            "Документы (*.pdf *.jpg *.jpeg *.png *.doc *.docx)"
        )
        
        if not file_path:
            return
            
        # Получаем имя файла
        file_name = os.path.basename(file_path)
        
        # Проверяем, существует ли папка для документов оборудования
        inv_number = self.inventory_number_input.text().strip()
        if not inv_number:
            QMessageBox.warning(self, "Ошибка", "Сначала введите инвентарный номер")
            return
            
        # Создаем структуру папок, если её нет
        doc_dir = os.path.join(DOCS_FOLDER, inv_number)
        os.makedirs(doc_dir, exist_ok=True)
        
        # Путь для сохранения копии файла
        dest_path = os.path.join(doc_dir, file_name)
        
        try:
            # Копируем файл в папку оборудования
            shutil.copy2(file_path, dest_path)
            
            # Добавляем документ в список
            doc_info = {
                "name": file_name,
                "path": os.path.join(inv_number, file_name),
                "date_added": datetime.now().strftime("%Y-%m-%d")
            }
            
            self.equipment["documents"].append(doc_info)
            self.docs_list.addItem(file_name)
            
            QMessageBox.information(self, "Успех", "Документ успешно прикреплен")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось прикрепить документ: {str(e)}")

    def remove_document(self):
        current_item = self.docs_list.currentRow()
        if current_item == -1:
            QMessageBox.warning(self, "Предупреждение", "Выберите документ для удаления")
            return
            
        doc_name = self.docs_list.currentItem().text()
        reply = QMessageBox.question(
            self, 
            "Подтверждение", 
            f"Вы уверены, что хотите удалить документ '{doc_name}'?",
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Удаляем документ из списка
            for i, doc in enumerate(self.equipment["documents"]):
                if doc["name"] == doc_name:
                    try:
                        # Удаляем файл с диска
                        doc_path = os.path.join(DOCS_FOLDER, doc["path"])
                        if os.path.exists(doc_path):
                            os.remove(doc_path)
                            
                        # Удаляем из списка документов
                        self.equipment["documents"].pop(i)
                        self.docs_list.takeItem(current_item)
                        
                        QMessageBox.information(self, "Успех", "Документ успешно удален")
                        break
                    except Exception as e:
                        QMessageBox.critical(self, "Ошибка", f"Не удалось удалить документ: {str(e)}")

    def validate_and_save(self):
        name = self.name_input.text().strip()
        inventory_number = self.inventory_number_input.text().strip()
        serial_number = self.serial_number_input.text().strip()

        if not name:
            QMessageBox.warning(self, "Ошибка", "Введите наименование оборудования")
            return

        if not inventory_number:
            QMessageBox.warning(self, "Ошибка", "Введите инвентарный номер")
            return

        verification_date = self.verification_date_edit.date().toString("yyyy-MM-dd")
        next_verification_date = self.next_verification_date_edit.date().toString(
            "yyyy-MM-dd"
        )

        # Генерируем уникальный ID для оборудования, если его нет
        if "equipment_id" not in self.equipment:
            self.equipment["equipment_id"] = f"EQ-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # Обновляем данные оборудования
        self.equipment.update({
            "name": name,
            "inventory_number": inventory_number,
            "serial_number": serial_number,
            "department": self.department_combo.currentText(),
            "verification_date": verification_date,
            "next_verification_date": next_verification_date,
            "notes": self.notes_input.toPlainText(),
        })

        self.accept()

class EquipmentDetailsDialog(QDialog):
    def __init__(self, equipment, parent=None):
        super().__init__(parent)
        self.equipment = equipment
        self.setWindowTitle(f"Карточка оборудования: {equipment['name']}")
        self.setMinimumSize(800, 600)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Вкладки для разделов информации
        tabs = QTabWidget()
        
        # Вкладка "Основная информация"
        info_tab = QWidget()
        info_layout = QFormLayout(info_tab)
        
        # Заголовок с наименованием оборудования
        title = QLabel(self.equipment["name"])
        title.setStyleSheet("font: bold 18pt 'Segoe UI';")
        layout.addWidget(title)
        
        # Основные поля
        fields = [
            ("ID оборудования:", self.equipment.get("equipment_id", "Не задан")),
            ("Инвентарный номер:", self.equipment.get("inventory_number", "")),
            ("Серийный номер:", self.equipment.get("serial_number", "")),
            ("Подразделение:", self.equipment.get("department", "")),
        ]
        
        for label_text, value in fields:
            label = QLabel(label_text)
            label.setStyleSheet("font-weight: bold;")
            value_label = QLabel(value)
            info_layout.addRow(label, value_label)
        
        # Даты поверок
        verif_date = datetime.strptime(
            self.equipment["verification_date"], "%Y-%m-%d"
        ).strftime("%d.%m.%Y")
        next_verif_date = datetime.strptime(
            self.equipment["next_verification_date"], "%Y-%m-%d"
        ).strftime("%d.%m.%Y")
        
        verif_label = QLabel("Дата поверки:")
        verif_label.setStyleSheet("font-weight: bold;")
        verif_value = QLabel(verif_date)
        info_layout.addRow(verif_label, verif_value)
        
        next_verif_label = QLabel("Следующая поверка:")
        next_verif_label.setStyleSheet("font-weight: bold;")
        next_verif_value = QLabel(next_verif_date)
        info_layout.addRow(next_verif_label, next_verif_value)
        
        # Статус поверки
        today = datetime.now().date()
        next_date = datetime.strptime(
            self.equipment["next_verification_date"], "%Y-%m-%d"
        ).date()
        days_left = (next_date - today).days
        
        if days_left < 0:
            status_text = f"Просрочено (на {abs(days_left)} дней)"
            status_color = "#e74c3c"
        elif days_left <= 30:
            status_text = f"Истекает через {days_left} дней"
            status_color = "#f39c12"
        else:
            status_text = f"Активно (осталось {days_left} дней)"
            status_color = "#2ecc71"
        
        status_label = QLabel("Статус поверки:")
        status_label.setStyleSheet("font-weight: bold;")
        status_value = QLabel(status_text)
        status_value.setStyleSheet(f"color: {status_color}; font-weight: bold;")
        info_layout.addRow(status_label, status_value)
        
        # Примечания
        notes_label = QLabel("Примечания:")
        notes_label.setStyleSheet("font-weight: bold;")
        notes_value = QTextEdit()
        notes_value.setPlainText(self.equipment.get("notes", ""))
        notes_value.setReadOnly(True)
        notes_value.setMaximumHeight(80)
        info_layout.addRow(notes_label, notes_value)
        
        tabs.addTab(info_tab, "Основная информация")
        
        # Вкладка "История поверок"
        history_tab = QWidget()
        history_layout = QVBoxLayout(history_tab)
        
        history_label = QLabel("История поверок:")
        history_label.setStyleSheet("font-weight: bold;")
        history_layout.addWidget(history_label)
        
        history_table = QTableWidget()
        history_table.setColumnCount(3)
        history_table.setHorizontalHeaderLabels(["Дата поверки", "Срок действия до", "Примечание"])
        history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Добавляем текущую поверку в историю
        history_table.setRowCount(1)
        history_table.setItem(0, 0, QTableWidgetItem(verif_date))
        history_table.setItem(0, 1, QTableWidgetItem(next_verif_date))
        history_table.setItem(0, 2, QTableWidgetItem("Текущая поверка"))
        
        # Если есть история поверок, добавляем их
        if "verification_history" in self.equipment:
            for i, hist in enumerate(self.equipment["verification_history"], 1):
                history_table.setRowCount(i + 1)
                history_table.setItem(i, 0, QTableWidgetItem(
                    datetime.strptime(hist["date"], "%Y-%m-%d").strftime("%d.%m.%Y"))
                )
                history_table.setItem(i, 1, QTableWidgetItem(
                    datetime.strptime(hist["valid_until"], "%Y-%m-%d").strftime("%d.%m.%Y"))
                )
                history_table.setItem(i, 2, QTableWidgetItem(hist.get("notes", "")))
        
        history_layout.addWidget(history_table)
        tabs.addTab(history_tab, "История поверок")
        
        # Вкладка "Документы"
        docs_tab = QWidget()
        docs_layout = QVBoxLayout(docs_tab)
        
        docs_label = QLabel("Прикрепленные документы:")
        docs_label.setStyleSheet("font-weight: bold;")
        docs_layout.addWidget(docs_label)
        
        self.docs_list = QListWidget()
        
        if "documents" in self.equipment and self.equipment["documents"]:
            for doc in self.equipment["documents"]:
                item = QListWidgetItem(doc["name"])
                item.setData(Qt.UserRole, doc["path"])
                self.docs_list.addItem(item)
        else:
            self.docs_list.addItem("Нет прикрепленных документов")
            
        docs_layout.addWidget(self.docs_list)
        
        view_doc_btn = QPushButton("Просмотреть документ")
        view_doc_btn.clicked.connect(self.view_document)
        docs_layout.addWidget(view_doc_btn)
        
        tabs.addTab(docs_tab, "Документы")
        
        layout.addWidget(tabs)
        
        # Кнопки внизу диалога
        btn_layout = QHBoxLayout()
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)

    def view_document(self):
        if not self.docs_list.currentItem():
            QMessageBox.warning(self, "Предупреждение", "Выберите документ для просмотра")
            return
            
        if self.docs_list.currentItem().text() == "Нет прикрепленных документов":
            return
            
        doc_path = self.docs_list.currentItem().data(Qt.UserRole)
        full_path = os.path.join(DOCS_FOLDER, doc_path)
        
        if not os.path.exists(full_path):
            QMessageBox.warning(self, "Ошибка", "Файл не найден")
            return
            
        # Открываем файл в системном приложении по умолчанию
        QDesktopServices.openUrl(QUrl.fromLocalFile(full_path))


class UserManagementDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Упр-ние пользователями")
        self.setFixedSize(600, 400)
        self.setup_ui()
        self.load_users()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Таблица пользователей
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)
        self.users_table.setHorizontalHeaderLabels(
            ["Логин", "Роль", "Email", "Действия"]
        )
        self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.users_table.verticalHeader().setVisible(False)
        self.users_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.users_table)

        # Кнопки управления
        btn_layout = QHBoxLayout()

        self.add_btn = QPushButton("Добавить пользователя")
        self.add_btn.clicked.connect(self.add_user)
        btn_layout.addWidget(self.add_btn)

        self.edit_btn = QPushButton("Редактировать")
        self.edit_btn.clicked.connect(self.edit_user)
        btn_layout.addWidget(self.edit_btn)

        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.clicked.connect(self.delete_user)
        btn_layout.addWidget(self.delete_btn)

        self.close_btn = QPushButton("Закрыть")
        self.close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.close_btn)

        layout.addLayout(btn_layout)

    def load_users(self):
        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            self.users_table.setRowCount(len(users))

            for row, (username, user_data) in enumerate(users.items()):
                # Логин
                login_item = QTableWidgetItem(username)
                login_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 0, login_item)

                # Роль
                role_item = QTableWidgetItem(
                    ROLES.get(user_data.get("role", "guest"), "Гость")
                )
                role_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 1, role_item)

                # Email
                email_item = QTableWidgetItem(user_data.get("email", ""))
                email_item.setTextAlignment(Qt.AlignCenter)
                self.users_table.setItem(row, 2, email_item)

                # Кнопка сброса пароля
                reset_btn = QPushButton("Сбросить пароль")
                reset_btn.clicked.connect(lambda _, u=username: self.reset_password(u))
                self.users_table.setCellWidget(row, 3, reset_btn)

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось загрузить пользователей: {str(e)}"
            )

    def add_user(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить пользователя")
        dialog.setFixedSize(400, 300)

        layout = QFormLayout(dialog)

        # Поля формы
        username_input = QLineEdit()
        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.Password)
        confirm_password_input = QLineEdit()
        confirm_password_input.setEchoMode(QLineEdit.Password)

        role_combo = QComboBox()
        for role_key, role_name in ROLES.items():
            role_combo.addItem(role_name, role_key)

        email_input = QLineEdit()

        # Добавление полей в форму
        layout.addRow("Логин:", username_input)
        layout.addRow("Пароль:", password_input)
        layout.addRow("Подтверждение пароля:", confirm_password_input)
        layout.addRow("Роль:", role_combo)
        layout.addRow("Email:", email_input)

        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(
            lambda: self.validate_and_add_user(
                dialog,
                username_input.text(),
                password_input.text(),
                confirm_password_input.text(),
                role_combo.currentData(),
                email_input.text(),
            )
        )
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)

        dialog.exec_()

    def validate_and_add_user(
        self, dialog, username, password, confirm_password, role, email
    ):
        if not username:
            QMessageBox.warning(self, "Ошибка", "Введите логин пользователя")
            return

        if not password:
            QMessageBox.warning(self, "Ошибка", "Введите пароль")
            return

        if password != confirm_password:
            QMessageBox.warning(self, "Ошибка", "Пароли не совпадают")
            return

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username in users:
                QMessageBox.warning(
                    self, "Ошибка", "Пользователь с таким логином уже существует"
                )
                return

            users[username] = {
                "password": sha256(password.encode("utf-8")).hexdigest(),
                "role": role,
                "email": email,
            }

            with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(users, f, indent=4, ensure_ascii=False)

            QMessageBox.information(self, "Успех", "Пользователь успешно добавлен")
            self.load_users()
            dialog.accept()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось добавить пользователя: {str(e)}"
            )

    def edit_user(self):
        selected_row = self.users_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выберите пользователя для редактирования"
            )
            return

        username = self.users_table.item(selected_row, 0).text()

        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username not in users:
                QMessageBox.warning(self, "Ошибка", "Выбранный пользователь не найден")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Редактировать пользователя")
            dialog.setFixedSize(400, 250)

            layout = QFormLayout(dialog)

            # Поля формы
            role_combo = QComboBox()
            for role_key, role_name in ROLES.items():
                role_combo.addItem(role_name, role_key)

            # Устанавливаем текущую роль
            current_role = users[username].get("role", "guest")
            index = role_combo.findData(current_role)
            if index >= 0:
                role_combo.setCurrentIndex(index)

            email_input = QLineEdit(users[username].get("email", ""))

            # Добавление полей в форму
            layout.addRow("Логин:", QLabel(username))
            layout.addRow("Роль:", role_combo)
            layout.addRow("Email:", email_input)

            # Кнопки
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            buttons.accepted.connect(
                lambda: self.save_user_changes(
                    dialog, username, role_combo.currentData(), email_input.text()
                )
            )
            buttons.rejected.connect(dialog.reject)
            layout.addRow(buttons)

            dialog.exec_()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось редактировать пользователя: {str(e)}"
            )

    def save_user_changes(self, dialog, username, role, email):
        try:
            with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)

            if username not in users:
                QMessageBox.warning(self, "Ошибка", "Пользователь не найден")
                return

            users[username]["role"] = role
            users[username]["email"] = email

            with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(users, f, indent=4, ensure_ascii=False)

            QMessageBox.information(self, "Успех", "Изменения сохранены")
            self.load_users()
            dialog.accept()

        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
            )

    def delete_user(self):
        selected_row = self.users_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите пользователя для удаления")
            return

        username = self.users_table.item(selected_row, 0).text()

        if username == "admin":
            QMessageBox.warning(
                self, "Ошибка", "Нельзя удалить администратора по умолчанию"
            )
            return

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить пользователя {username}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                    users = json.load(f)

                if username in users:
                    del users[username]

                    with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                        json.dump(users, f, indent=4, ensure_ascii=False)

                    QMessageBox.information(self, "Успех", "Пользователь удален")
                    self.load_users()

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить пользователя: {str(e)}"
                )

    def reset_password(self, username):
        if username == "admin":
            QMessageBox.warning(
                self,
                "Ошибка",
                "Для сброса пароля администратора используйте функцию изменения пароля",
            )
            return

        new_password, ok = QInputDialog.getText(
            self,
            "Сброс пароля",
            f"Введите новый пароль для пользователя {username}:",
            QLineEdit.Password,
        )

        if ok and new_password:
            try:
                with open(USER_DATA_FILE, "r", encoding="utf-8") as f:
                    users = json.load(f)

                if username in users:
                    users[username]["password"] = sha256(
                        new_password.encode("utf-8")
                    ).hexdigest()

                    with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
                        json.dump(users, f, indent=4, ensure_ascii=False)

                    QMessageBox.information(self, "Успех", "Пароль успешно изменен")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Ошибка", "Пользователь не найден")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось изменить пароль: {str(e)}"
                )

class SpecialistDialog(QDialog):
    def __init__(self, specialist=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(
            "Добавить специалиста" if specialist is None else "Редактировать специалиста"
        )
        self.setFixedSize(600, 500)
        self.specialist = specialist or {}
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Форма для ввода данных
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # Поля формы
        self.last_name_input = QLineEdit()
        form_layout.addRow("Фамилия:", self.last_name_input)

        self.first_name_input = QLineEdit()
        form_layout.addRow("Имя:", self.first_name_input)

        self.middle_name_input = QLineEdit()
        form_layout.addRow("Отчество:", self.middle_name_input)

        self.position_input = QLineEdit()
        form_layout.addRow("Должность:", self.position_input)

        self.department_combo = QComboBox()
        self.department_combo.addItems(DEPARTMENTS)
        form_layout.addRow("Подразделение:", self.department_combo)

        # Методы контроля
        self.methods_group = QFrame()
        methods_layout = QVBoxLayout(self.methods_group)
        methods_label = QLabel("Методы контроля:")
        methods_layout.addWidget(methods_label)

        # Чекбоксы для методов контроля
        self.methods_checkboxes = {}
        methods_grid = QGridLayout()
        for i, method in enumerate(CONTROL_METHODS):
            row, col = divmod(i, 5)
            cb = QComboBox()
            cb.addItems(["Нет", "I уровень", "II уровень", "III уровень"])
            self.methods_checkboxes[method] = cb
            methods_grid.addWidget(QLabel(method), row, col*2)
            methods_grid.addWidget(cb, row, col*2+1)
        
        methods_layout.addLayout(methods_grid)
        form_layout.addRow("", self.methods_group)

        # Сертификаты
        self.cert_number_input = QLineEdit()
        form_layout.addRow("Номер удостоверения:", self.cert_number_input)

        self.cert_date_edit = QDateEdit()
        self.cert_date_edit.setCalendarPopup(True)
        self.cert_date_edit.setDate(datetime.now().date())
        form_layout.addRow("Дата аттестации:", self.cert_date_edit)

        self.next_cert_date_edit = QDateEdit()
        self.next_cert_date_edit.setCalendarPopup(True)
        self.next_cert_date_edit.setDate(
            datetime.now().date() + timedelta(days=365)
        )
        form_layout.addRow("Срок действия до:", self.next_cert_date_edit)

        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(60)
        self.notes_input.setPlaceholderText("Дополнительная информация")
        form_layout.addRow("Примечания:", self.notes_input)

        layout.addLayout(form_layout)

        # Если редактируем существующего специалиста, заполняем поля
        if self.specialist:
            self.last_name_input.setText(self.specialist.get("last_name", ""))
            self.first_name_input.setText(self.specialist.get("first_name", ""))
            self.middle_name_input.setText(self.specialist.get("middle_name", ""))
            self.position_input.setText(self.specialist.get("position", ""))
            self.department_combo.setCurrentText(
                self.specialist.get("department", DEPARTMENTS[0])
            )
            
            # Заполняем методы контроля
            methods = self.specialist.get("methods", {})
            for method, level in methods.items():
                if method in self.methods_checkboxes:
                    if level == 1:
                        self.methods_checkboxes[method].setCurrentIndex(1)
                    elif level == 2:
                        self.methods_checkboxes[method].setCurrentIndex(2)
                    elif level == 3:
                        self.methods_checkboxes[method].setCurrentIndex(3)
            
            self.cert_number_input.setText(self.specialist.get("cert_number", ""))
            
            if "cert_date" in self.specialist:
                self.cert_date_edit.setDate(
                    datetime.strptime(
                        self.specialist["cert_date"], "%Y-%m-%d"
                    ).date()
                )
                
            if "next_cert_date" in self.specialist:
                self.next_cert_date_edit.setDate(
                    datetime.strptime(
                        self.specialist["next_cert_date"], "%Y-%m-%d"
                    ).date()
                )
                
            self.notes_input.setPlainText(self.specialist.get("notes", ""))

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_btn.clicked.connect(self.validate_and_save)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def validate_and_save(self):
        last_name = self.last_name_input.text().strip()
        first_name = self.first_name_input.text().strip()
        cert_number = self.cert_number_input.text().strip()

        if not last_name or not first_name:
            QMessageBox.warning(self, "Ошибка", "Введите фамилию и имя специалиста")
            return

        if not cert_number:
            QMessageBox.warning(self, "Ошибка", "Введите номер удостоверения")
            return

        # Собираем методы контроля
        methods = {}
        for method, combo in self.methods_checkboxes.items():
            level_index = combo.currentIndex()
            if level_index > 0:  # Если выбран уровень
                methods[method] = level_index

        # Генерируем уникальный ID для специалиста, если его нет
        if "specialist_id" not in self.specialist:
            self.specialist["specialist_id"] = f"SP-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # Обновляем данные специалиста
        self.specialist.update({
            "last_name": last_name,
            "first_name": first_name,
            "middle_name": self.middle_name_input.text().strip(),
            "full_name": f"{last_name} {first_name} {self.middle_name_input.text().strip()}",
            "position": self.position_input.text().strip(),
            "department": self.department_combo.currentText(),
            "methods": methods,
            "cert_number": cert_number,
            "cert_date": self.cert_date_edit.date().toString("yyyy-MM-dd"),
            "next_cert_date": self.next_cert_date_edit.date().toString("yyyy-MM-dd"),
            "notes": self.notes_input.toPlainText(),
        })

        self.accept()

class SpecialistsWidget(QWidget):
    def __init__(self, parent=None, user_role="guest"):
        super().__init__(parent)
        self.user_role = user_role
        self.specialists_data = []
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_layout = QHBoxLayout()
        title = QLabel("Управление специалистами")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        header_layout.addWidget(title)

        # Кнопки управления (только для администратора, метролога и руководителя)
        if self.user_role in ["admin", "metrolog", "manager"]:
            btn_layout = QHBoxLayout()
            btn_layout.setSpacing(10)

            add_btn = QPushButton("Добавить")
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            add_btn.clicked.connect(self.add_specialist)
            btn_layout.addWidget(add_btn)

            edit_btn = QPushButton("Редактировать")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #0b7dda;
                }
            """)
            edit_btn.clicked.connect(self.edit_specialist)
            btn_layout.addWidget(edit_btn)

            delete_btn = QPushButton("Удалить")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #d32f2f;
                }
            """)
            delete_btn.clicked.connect(self.delete_specialist)
            btn_layout.addWidget(delete_btn)

            header_layout.addLayout(btn_layout)

        layout.addLayout(header_layout)

        # Фильтры
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.department_filter = QComboBox()
        self.department_filter.addItem("Все подразделения")
        for dept in DEPARTMENTS:
            self.department_filter.addItem(dept)
        self.department_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Подразделение:"))
        filter_layout.addWidget(self.department_filter)

        self.method_filter = QComboBox()
        self.method_filter.addItem("Все методы", "")
        for method in CONTROL_METHODS:
            self.method_filter.addItem(method, method)
        self.method_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Метод:"))
        filter_layout.addWidget(self.method_filter)

        self.status_filter = QComboBox()
        self.status_filter.addItem("Все статусы", "")
        self.status_filter.addItem("Активные", "active")
        self.status_filter.addItem("Истекающие", "expiring")
        self.status_filter.addItem("Просроченные", "overdue")
        self.status_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Статус:"))
        filter_layout.addWidget(self.status_filter)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по ФИО...")
        self.search_input.textChanged.connect(self.apply_filters)
        search_layout.addWidget(self.search_input)

        export_btn = QPushButton("Экспорт")
        export_btn.clicked.connect(self.export_specialists)
        search_layout.addWidget(export_btn)

        filter_layout.addLayout(search_layout)
        layout.addLayout(filter_layout)

        # Таблица специалистов
        self.specialists_table = QTableWidget()
        self.specialists_table.setColumnCount(6)
        self.specialists_table.setHorizontalHeaderLabels(
            [
                "ФИО",
                "Должность",
                "Подразделение",
                "Методы контроля",
                "Срок действия до",
                "Статус",
            ]
        )
        self.specialists_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.specialists_table.verticalHeader().setVisible(False)
        self.specialists_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.specialists_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.specialists_table.setSelectionMode(QTableWidget.SingleSelection)
        self.specialists_table.setSortingEnabled(True)
        self.specialists_table.doubleClicked.connect(self.view_specialist_details)
        layout.addWidget(self.specialists_table)

    def load_data(self):
        try:
            if not os.path.exists(SPECIALISTS_DATA_FILE):
                with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump([], f)

            with open(SPECIALISTS_DATA_FILE, "r", encoding="utf-8") as f:
                self.specialists_data = json.load(f)

            self.update_specialists_table()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка загрузки",
                f"Не удалось загрузить данные специалистов: {str(e)}",
            )

    def update_specialists_table(self):
        self.specialists_table.setRowCount(len(self.specialists_data))

        for row, specialist in enumerate(self.specialists_data):
            # ФИО
            name_item = QTableWidgetItem(specialist.get("full_name", ""))
            name_item.setData(Qt.UserRole, specialist.get("specialist_id", ""))
            self.specialists_table.setItem(row, 0, name_item)

            # Должность
            position_item = QTableWidgetItem(specialist.get("position", ""))
            self.specialists_table.setItem(row, 1, position_item)

            # Подразделение
            dept_item = QTableWidgetItem(specialist.get("department", ""))
            self.specialists_table.setItem(row, 2, dept_item)

            # Методы контроля
            methods = specialist.get("methods", {})
            methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
            methods_item = QTableWidgetItem(methods_text)
            self.specialists_table.setItem(row, 3, methods_item)

            # Срок действия
            if "next_cert_date" in specialist:
                next_date = datetime.strptime(
                    specialist["next_cert_date"], "%Y-%m-%d"
                ).date()
                next_date_item = QTableWidgetItem(next_date.strftime("%d.%m.%Y"))
                self.specialists_table.setItem(row, 4, next_date_item)
            else:
                self.specialists_table.setItem(row, 4, QTableWidgetItem("Не указан"))

            # Статус
            if "next_cert_date" in specialist:
                today = datetime.now().date()
                next_date = datetime.strptime(
                    specialist["next_cert_date"], "%Y-%m-%d"
                ).date()
                days_left = (next_date - today).days

                if days_left < 0:
                    status_text = "Просрочено"
                    status_color = "#e74c3c"
                elif days_left <= 30:
                    status_text = f"Истекает ({days_left} дн.)"
                    status_color = "#f39c12"
                else:
                    status_text = f"Активно ({days_left} дн.)"
                    status_color = "#2ecc71"

                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QColor(status_color))
                self.specialists_table.setItem(row, 5, status_item)
            else:
                status_item = QTableWidgetItem("Не указан")
                self.specialists_table.setItem(row, 5, status_item)

    def apply_filters(self):
        department_filter = self.department_filter.currentText()
        method_filter = self.method_filter.currentData()
        status_filter = self.status_filter.currentData()
        search_text = self.search_input.text().lower()
        today = datetime.now().date()

        for row in range(self.specialists_table.rowCount()):
            should_show = True

            # Фильтр по подразделению
            if department_filter != "Все подразделения":
                dept_item = self.specialists_table.item(row, 2)
                if dept_item.text() != department_filter:
                    should_show = False

            # Фильтр по методу контроля
            if method_filter and should_show:
                methods_item = self.specialists_table.item(row, 3)
                if method_filter not in methods_item.text():
                    should_show = False

            # Фильтр по статусу
            if status_filter and should_show:
                next_date_item = self.specialists_table.item(row, 4)
                if next_date_item.text() != "Не указан":
                    next_date = datetime.strptime(
                        next_date_item.text(), "%d.%m.%Y"
                    ).date()
                    days_left = (next_date - today).days

                    if status_filter == "active" and days_left <= 30:
                        should_show = False
                    elif status_filter == "expiring" and not (0 <= days_left <= 30):
                        should_show = False
                    elif status_filter == "overdue" and days_left >= 0:
                        should_show = False

            # Фильтр по поиску
            if should_show and search_text:
                name_item = self.specialists_table.item(row, 0)
                if search_text not in name_item.text().lower():
                    should_show = False

            self.specialists_table.setRowHidden(row, not should_show)

    def add_specialist(self):
        dialog = SpecialistDialog()
        if dialog.exec_() == QDialog.Accepted:
            try:
                self.specialists_data.append(dialog.specialist)

                with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.specialists_data, f, indent=4, ensure_ascii=False)

                self.update_specialists_table()
                QMessageBox.information(self, "Успех", "Специалист успешно добавлен")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось добавить специалиста: {str(e)}"
                )

    def edit_specialist(self):
        selected_row = self.specialists_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выберите специалиста для редактирования"
            )
            return

        specialist_id = self.specialists_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующего специалиста в массиве данных
        found_index = -1
        for i, spec in enumerate(self.specialists_data):
            if spec.get("specialist_id") == specialist_id:
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранный специалист не найден в базе данных"
            )
            return

        specialist = self.specialists_data[found_index]
        dialog = SpecialistDialog(specialist, self)

        if dialog.exec_() == QDialog.Accepted:
            try:
                self.specialists_data[found_index] = dialog.specialist

                with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.specialists_data, f, indent=4, ensure_ascii=False)

                self.update_specialists_table()
                QMessageBox.information(self, "Успех", "Изменения сохранены")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
                )

    def delete_specialist(self):
        selected_row = self.specialists_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите специалиста для удаления")
            return

        specialist_id = self.specialists_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующего специалиста в массиве данных
        found_index = -1
        for i, spec in enumerate(self.specialists_data):
            if spec.get("specialist_id") == specialist_id:
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранный специалист не найден в базе данных"
            )
            return

        specialist = self.specialists_data[found_index]

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить специалиста {specialist.get('full_name', '')}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                del self.specialists_data[found_index]

                with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.specialists_data, f, indent=4, ensure_ascii=False)

                self.update_specialists_table()
                QMessageBox.information(self, "Успех", "Специалист удален")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить специалиста: {str(e)}"
                )

    def view_specialist_details(self):
        selected_row = self.specialists_table.currentRow()
        if selected_row == -1:
            return

        specialist_id = self.specialists_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующего специалиста в массиве данных
        found_specialist = None
        for spec in self.specialists_data:
            if spec.get("specialist_id") == specialist_id:
                found_specialist = spec
                break

        if not found_specialist:
            return

        # Создаем диалог с подробной информацией о специалисте
        details_dialog = QDialog(self)
        details_dialog.setWindowTitle(f"Информация о специалисте: {found_specialist.get('full_name', '')}")
        details_dialog.setMinimumSize(600, 400)

        layout = QVBoxLayout(details_dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # Личные данные
        personal_group = QGroupBox("Личные данные")
        personal_layout = QFormLayout(personal_group)

        personal_layout.addRow("ФИО:", QLabel(found_specialist.get("full_name", "")))
        personal_layout.addRow("Должность:", QLabel(found_specialist.get("position", "")))
        personal_layout.addRow("Подразделение:", QLabel(found_specialist.get("department", "")))

        layout.addWidget(personal_group)

        # Сертификация
        cert_group = QGroupBox("Сертификация")
        cert_layout = QFormLayout(cert_group)

        cert_layout.addRow("Номер удостоверения:", QLabel(found_specialist.get("cert_number", "")))
        
        if "cert_date" in found_specialist:
            cert_date = datetime.strptime(found_specialist["cert_date"], "%Y-%m-%d").strftime("%d.%m.%Y")
            cert_layout.addRow("Дата аттестации:", QLabel(cert_date))
            
        if "next_cert_date" in found_specialist:
            next_date = datetime.strptime(found_specialist["next_cert_date"], "%Y-%m-%d").date()
            next_date_text = next_date.strftime("%d.%m.%Y")
            
            today = datetime.now().date()
            days_left = (next_date - today).days
            
            status_label = QLabel()
            if days_left < 0:
                status_text = f"Просрочено (на {abs(days_left)} дней)"
                status_label.setStyleSheet("color: #e74c3c; font-weight: bold;")
            elif days_left <= 30:
                status_text = f"Истекает через {days_left} дней"
                status_label.setStyleSheet("color: #f39c12; font-weight: bold;")
            else:
                status_text = f"Активно (осталось {days_left} дней)"
                status_label.setStyleSheet("color: #2ecc71; font-weight: bold;")
                
            status_label.setText(status_text)
            
            cert_layout.addRow("Срок действия до:", QLabel(next_date_text))
            cert_layout.addRow("Статус:", status_label)

        layout.addWidget(cert_group)

        # Методы контроля
        methods_group = QGroupBox("Методы контроля")
        methods_layout = QVBoxLayout(methods_group)

        methods = found_specialist.get("methods", {})
        if methods:
            methods_table = QTableWidget()
            methods_table.setColumnCount(2)
            methods_table.setHorizontalHeaderLabels(["Метод", "Уровень"])
            methods_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            methods_table.setRowCount(len(methods))
            
            for i, (method, level) in enumerate(methods.items()):
                methods_table.setItem(i, 0, QTableWidgetItem(method))
                methods_table.setItem(i, 1, QTableWidgetItem(str(level)))
                
            methods_layout.addWidget(methods_table)
        else:
            methods_layout.addWidget(QLabel("Нет аттестованных методов контроля"))

        layout.addWidget(methods_group)

        # Примечания
        if found_specialist.get("notes"):
            notes_group = QGroupBox("Примечания")
            notes_layout = QVBoxLayout(notes_group)
            notes_text = QTextEdit()
            notes_text.setPlainText(found_specialist.get("notes", ""))
            notes_text.setReadOnly(True)
            notes_layout.addWidget(notes_text)
            layout.addWidget(notes_group)

        # Кнопка закрытия
        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(details_dialog.accept)
        layout.addWidget(close_button)

        details_dialog.exec_()

    def export_specialists(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт данных",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
            options=options,
        )

        if not file_name:
            return

        try:
            if file_name.endswith(".xlsx"):
                self.export_to_excel(file_name)
            else:
                self.export_to_csv(file_name)

            QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось экспортировать данные: {str(e)}"
            )

    def export_to_excel(self, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Специалисты"

        # Заголовки
        headers = [
            "ФИО",
            "Должность",
            "Подразделение",
            "Номер удостоверения",
            "Дата аттестации",
            "Срок действия до",
            "Статус",
            "Методы контроля",
            "Примечания",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Данные
        today = datetime.now().date()

        for row, specialist in enumerate(self.specialists_data, 2):
            sheet.cell(row=row, column=1, value=specialist.get("full_name", ""))
            sheet.cell(row=row, column=2, value=specialist.get("position", ""))
            sheet.cell(row=row, column=3, value=specialist.get("department", ""))
            sheet.cell(row=row, column=4, value=specialist.get("cert_number", ""))
            
            if "cert_date" in specialist:
                cert_date = datetime.strptime(specialist["cert_date"], "%Y-%m-%d").date()
                sheet.cell(row=row, column=5, value=cert_date.strftime("%d.%m.%Y"))
            
            if "next_cert_date" in specialist:
                next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                sheet.cell(row=row, column=6, value=next_date.strftime("%d.%m.%Y"))
                
                days_left = (next_date - today).days
                if days_left < 0:
                    status = f"Просрочено (на {abs(days_left)} дней)"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                else:
                    status = f"Активно (осталось {days_left} дней)"
                    
                sheet.cell(row=row, column=7, value=status)
            
            methods = specialist.get("methods", {})
            methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
            sheet.cell(row=row, column=8, value=methods_text)
            
            sheet.cell(row=row, column=9, value=specialist.get("notes", ""))

        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Границы для всех ячеек
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        workbook.save(file_path)

    def export_to_csv(self, file_path):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            # Заголовки
            writer.writerow(
                [
                    "ФИО",
                    "Должность",
                    "Подразделение",
                    "Номер удостоверения",
                    "Дата аттестации",
                    "Срок действия до",
                    "Статус",
                    "Методы контроля",
                    "Примечания",
                ]
            )

            # Данные
            today = datetime.now().date()

            for specialist in self.specialists_data:
                status = ""
                if "next_cert_date" in specialist:
                    next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                    days_left = (next_date - today).days
                    
                    if days_left < 0:
                        status = f"Просрочено (на {abs(days_left)} дней)"
                    elif days_left <= 30:
                        status = f"Истекает через {days_left} дней"
                    else:
                        status = f"Активно (осталось {days_left} дней)"
                        
                methods = specialist.get("methods", {})
                methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
                
                writer.writerow(
                    [
                        specialist.get("full_name", ""),
                        specialist.get("position", ""),
                        specialist.get("department", ""),
                        specialist.get("cert_number", ""),
                        datetime.strptime(specialist["cert_date"], "%Y-%m-%d").strftime("%d.%m.%Y") if "cert_date" in specialist else "",
                        datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").strftime("%d.%m.%Y") if "next_cert_date" in specialist else "",
                        status,
                        methods_text,
                        specialist.get("notes", ""),
                    ]
                )

class ObjectDialog(QDialog):
    def __init__(self, object_data=None, equipment_data=[], specialists_data=[], parent=None):
        super().__init__(parent)
        self.setWindowTitle(
            "Добавить объект" if object_data is None else "Редактировать объект"
        )
        self.setFixedSize(700, 600)
        self.object_data = object_data or {}
        self.equipment_data = equipment_data
        self.specialists_data = specialists_data
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Вкладки
        tabs = QTabWidget()
        
        # Вкладка "Основная информация"
        main_tab = QWidget()
        main_layout = QFormLayout(main_tab)
        
        # Поля формы
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Название объекта контроля")
        main_layout.addRow("Наименование:", self.name_input)
        
        self.number_input = QLineEdit()
        self.number_input.setPlaceholderText("Номер объекта/заказа")
        main_layout.addRow("Номер:", self.number_input)
        
        self.department_combo = QComboBox()
        self.department_combo.addItems(DEPARTMENTS)
        main_layout.addRow("Подразделение:", self.department_combo)
        
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(datetime.now().date())
        main_layout.addRow("Дата создания:", self.date_edit)
        
        self.customer_input = QLineEdit()
        self.customer_input.setPlaceholderText("Наименование заказчика")
        main_layout.addRow("Заказчик:", self.customer_input)
        
        self.address_input = QLineEdit()
        self.address_input.setPlaceholderText("Адрес объекта")
        main_layout.addRow("Адрес:", self.address_input)
        
        self.notes_input = QTextEdit()
        self.notes_input.setMaximumHeight(80)
        self.notes_input.setPlaceholderText("Дополнительная информация")
        main_layout.addRow("Примечания:", self.notes_input)
        
        tabs.addTab(main_tab, "Основная информация")
        
        # Вкладка "Оборудование"
        equipment_tab = QWidget()
        equipment_layout = QVBoxLayout(equipment_tab)
        
        equipment_label = QLabel("Задействованное оборудование:")
        equipment_layout.addWidget(equipment_label)
        
        self.equipment_list = QListWidget()
        self.equipment_list.setSelectionMode(QListWidget.MultiSelection)
        
        # Заполняем список доступного оборудования
        for eq in self.equipment_data:
            item = QListWidgetItem(f"{eq['name']} (инв. № {eq['inventory_number']})")
            item.setData(Qt.UserRole, eq.get('equipment_id', ''))
            self.equipment_list.addItem(item)
            
        equipment_layout.addWidget(self.equipment_list)
        
        tabs.addTab(equipment_tab, "Оборудование")
        
        # Вкладка "Специалисты"
        specialists_tab = QWidget()
        specialists_layout = QVBoxLayout(specialists_tab)
        
        specialists_label = QLabel("Ответственные специалисты:")
        specialists_layout.addWidget(specialists_label)
        
        self.specialists_list = QListWidget()
        self.specialists_list.setSelectionMode(QListWidget.MultiSelection)
        
        # Заполняем список доступных специалистов
        for spec in self.specialists_data:
            item = QListWidgetItem(f"{spec.get('full_name', '')} ({spec.get('position', '')})")
            item.setData(Qt.UserRole, spec.get('specialist_id', ''))
            self.specialists_list.addItem(item)
            
        specialists_layout.addWidget(self.specialists_list)
        
        tabs.addTab(specialists_tab, "Специалисты")
        
        # Вкладка "Заключение"
        conclusion_tab = QWidget()
        conclusion_layout = QFormLayout(conclusion_tab)
        
        self.conclusion_number_input = QLineEdit()
        self.conclusion_number_input.setPlaceholderText("Номер заключения")
        conclusion_layout.addRow("Номер заключения:", self.conclusion_number_input)
        
        self.conclusion_date_edit = QDateEdit()
        self.conclusion_date_edit.setCalendarPopup(True)
        self.conclusion_date_edit.setDate(datetime.now().date())
        conclusion_layout.addRow("Дата заключения:", self.conclusion_date_edit)
        
        self.conclusion_result_combo = QComboBox()
        self.conclusion_result_combo.addItems(["Годен", "Не годен", "Требуется доработка"])
        conclusion_layout.addRow("Результат:", self.conclusion_result_combo)
        
        self.conclusion_text = QTextEdit()
        self.conclusion_text.setPlaceholderText("Текст заключения")
        conclusion_layout.addRow("Текст заключения:", self.conclusion_text)
        
        tabs.addTab(conclusion_tab, "Заключение")
        
        layout.addWidget(tabs)
        
        # Если редактируем существующий объект, заполняем поля
        if self.object_data:
            self.name_input.setText(self.object_data.get("name", ""))
            self.number_input.setText(self.object_data.get("number", ""))
            self.department_combo.setCurrentText(self.object_data.get("department", DEPARTMENTS[0]))
            
            if "date" in self.object_data:
                self.date_edit.setDate(
                    datetime.strptime(self.object_data["date"], "%Y-%m-%d").date()
                )
                
            self.customer_input.setText(self.object_data.get("customer", ""))
            self.address_input.setText(self.object_data.get("address", ""))
            self.notes_input.setPlainText(self.object_data.get("notes", ""))
            
            # Выделяем выбранное оборудование
            if "equipment" in self.object_data:
                for i in range(self.equipment_list.count()):
                    item = self.equipment_list.item(i)
                    if item.data(Qt.UserRole) in self.object_data["equipment"]:
                        item.setSelected(True)
                        
            # Выделяем выбранных специалистов
            if "specialists" in self.object_data:
                for i in range(self.specialists_list.count()):
                    item = self.specialists_list.item(i)
                    if item.data(Qt.UserRole) in self.object_data["specialists"]:
                        item.setSelected(True)
                        
            # Заполняем данные заключения
            conclusion = self.object_data.get("conclusion", {})
            if conclusion:
                self.conclusion_number_input.setText(conclusion.get("number", ""))
                
                if "date" in conclusion:
                    self.conclusion_date_edit.setDate(
                        datetime.strptime(conclusion["date"], "%Y-%m-%d").date()
                    )
                    
                result_index = self.conclusion_result_combo.findText(conclusion.get("result", ""))
                if result_index >= 0:
                    self.conclusion_result_combo.setCurrentIndex(result_index)
                    
                self.conclusion_text.setPlainText(conclusion.get("text", ""))

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        save_btn.clicked.connect(self.validate_and_save)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def validate_and_save(self):
        name = self.name_input.text().strip()
        number = self.number_input.text().strip()

        if not name:
            QMessageBox.warning(self, "Ошибка", "Введите наименование объекта")
            return

        if not number:
            QMessageBox.warning(self, "Ошибка", "Введите номер объекта")
            return

        # Собираем выбранное оборудование
        selected_equipment = []
        for i in range(self.equipment_list.count()):
            item = self.equipment_list.item(i)
            if item.isSelected():
                selected_equipment.append(item.data(Qt.UserRole))
                
        # Собираем выбранных специалистов
        selected_specialists = []
        for i in range(self.specialists_list.count()):
            item = self.specialists_list.item(i)
            if item.isSelected():
                selected_specialists.append(item.data(Qt.UserRole))

        # Генерируем уникальный ID для объекта, если его нет
        if "object_id" not in self.object_data:
            self.object_data["object_id"] = f"OBJ-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # Данные заключения
        conclusion = {
            "number": self.conclusion_number_input.text().strip(),
            "date": self.conclusion_date_edit.date().toString("yyyy-MM-dd"),
            "result": self.conclusion_result_combo.currentText(),
            "text": self.conclusion_text.toPlainText()
        }

        # Обновляем данные объекта
        self.object_data.update({
            "name": name,
            "number": number,
            "department": self.department_combo.currentText(),
            "date": self.date_edit.date().toString("yyyy-MM-dd"),
            "customer": self.customer_input.text().strip(),
            "address": self.address_input.text().strip(),
            "notes": self.notes_input.toPlainText(),
            "equipment": selected_equipment,
            "specialists": selected_specialists,
            "conclusion": conclusion
        })

        self.accept()

class ObjectsWidget(QWidget):
    def __init__(self, parent=None, user_role="guest"):
        super().__init__(parent)
        self.user_role = user_role
        self.objects_data = []
        self.equipment_data = []
        self.specialists_data = []
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_layout = QHBoxLayout()
        title = QLabel("Журнал объектов и заключений")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        header_layout.addWidget(title)

        # Кнопки управления (только для определенных ролей)
        if self.user_role in ["admin", "expert", "manager"]:
            btn_layout = QHBoxLayout()
            btn_layout.setSpacing(10)

            add_btn = QPushButton("Добавить объект")
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            add_btn.clicked.connect(self.add_object)
            btn_layout.addWidget(add_btn)

            edit_btn = QPushButton("Редактировать")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #0b7dda;
                }
            """)
            edit_btn.clicked.connect(self.edit_object)
            btn_layout.addWidget(edit_btn)

            generate_btn = QPushButton("Заключение")
            generate_btn.setStyleSheet("""
                QPushButton {
                    background-color: #9c27b0;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #7b1fa2;
                }
            """)
            generate_btn.clicked.connect(self.generate_conclusion)
            btn_layout.addWidget(generate_btn)

            if self.user_role in ["admin", "manager"]:
                delete_btn = QPushButton("Удалить")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        padding: 8px 15px;
                        font: bold 12pt 'Segoe UI';
                    }
                    QPushButton:hover {
                        background-color: #d32f2f;
                    }
                """)
                delete_btn.clicked.connect(self.delete_object)
                btn_layout.addWidget(delete_btn)

            header_layout.addLayout(btn_layout)

        layout.addLayout(header_layout)

        # Фильтры
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.department_filter = QComboBox()
        self.department_filter.addItem("Все подразделения")
        for dept in DEPARTMENTS:
            self.department_filter.addItem(dept)
        self.department_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Подразделение:"))
        filter_layout.addWidget(self.department_filter)

        self.result_filter = QComboBox()
        self.result_filter.addItem("Все результаты", "")
        self.result_filter.addItem("Годен", "Годен")
        self.result_filter.addItem("Не годен", "Не годен")
        self.result_filter.addItem("Требуется доработка", "Требуется доработка")
        self.result_filter.addItem("Без заключения", "no_conclusion")
        self.result_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Результат:"))
        filter_layout.addWidget(self.result_filter)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по наименованию или номеру...")
        self.search_input.textChanged.connect(self.apply_filters)
        search_layout.addWidget(self.search_input)

        export_btn = QPushButton("Экспорт")
        export_btn.clicked.connect(self.export_objects)
        search_layout.addWidget(export_btn)

        filter_layout.addLayout(search_layout)
        layout.addLayout(filter_layout)

        # Таблица объектов
        self.objects_table = QTableWidget()
        self.objects_table.setColumnCount(7)
        self.objects_table.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Номер",
                "Подразделение",
                "Дата",
                "Заказчик",
                "Номер заключения",
                "Результат",
            ]
        )
        self.objects_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.objects_table.verticalHeader().setVisible(False)
        self.objects_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.objects_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.objects_table.setSelectionMode(QTableWidget.SingleSelection)
        self.objects_table.setSortingEnabled(True)
        self.objects_table.doubleClicked.connect(self.view_object_details)
        layout.addWidget(self.objects_table)

    def load_data(self):
        try:
            # Загружаем данные объектов
            if not os.path.exists(OBJECTS_DATA_FILE):
                with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump([], f)

            with open(OBJECTS_DATA_FILE, "r", encoding="utf-8") as f:
                self.objects_data = json.load(f)
                
            # Загружаем данные оборудования
            if os.path.exists(EQUIPMENT_DATA_FILE):
                with open(EQUIPMENT_DATA_FILE, "r", encoding="utf-8") as f:
                    self.equipment_data = json.load(f)
                    
            # Загружаем данные специалистов
            if os.path.exists(SPECIALISTS_DATA_FILE):
                with open(SPECIALISTS_DATA_FILE, "r", encoding="utf-8") as f:
                    self.specialists_data = json.load(f)

            self.update_objects_table()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка загрузки",
                f"Не удалось загрузить данные объектов: {str(e)}",
            )

    def update_objects_table(self):
        self.objects_table.setRowCount(len(self.objects_data))

        for row, obj in enumerate(self.objects_data):
            # Наименование
            name_item = QTableWidgetItem(obj.get("name", ""))
            name_item.setData(Qt.UserRole, obj.get("object_id", ""))
            self.objects_table.setItem(row, 0, name_item)

            # Номер
            number_item = QTableWidgetItem(obj.get("number", ""))
            self.objects_table.setItem(row, 1, number_item)

            # Подразделение
            dept_item = QTableWidgetItem(obj.get("department", ""))
            self.objects_table.setItem(row, 2, dept_item)

            # Дата
            if "date" in obj:
                date_obj = datetime.strptime(obj["date"], "%Y-%m-%d").date()
                date_item = QTableWidgetItem(date_obj.strftime("%d.%m.%Y"))
                self.objects_table.setItem(row, 3, date_item)
            else:
                self.objects_table.setItem(row, 3, QTableWidgetItem(""))

            # Заказчик
            customer_item = QTableWidgetItem(obj.get("customer", ""))
            self.objects_table.setItem(row, 4, customer_item)

            # Номер заключения и результат
            conclusion = obj.get("conclusion", {})
            if conclusion:
                conclusion_number = QTableWidgetItem(conclusion.get("number", ""))
                self.objects_table.setItem(row, 5, conclusion_number)
                
                result_item = QTableWidgetItem(conclusion.get("result", ""))
                # Цветовое оформление результата
                if conclusion.get("result") == "Годен":
                    result_item.setForeground(QColor("#2ecc71"))
                elif conclusion.get("result") == "Не годен":
                    result_item.setForeground(QColor("#e74c3c"))
                elif conclusion.get("result") == "Требуется доработка":
                    result_item.setForeground(QColor("#f39c12"))
                    
                self.objects_table.setItem(row, 6, result_item)
            else:
                self.objects_table.setItem(row, 5, QTableWidgetItem(""))
                no_conclusion = QTableWidgetItem("Нет заключения")
                no_conclusion.setForeground(QColor("#7f8c8d"))
                self.objects_table.setItem(row, 6, no_conclusion)

    def apply_filters(self):
        department_filter = self.department_filter.currentText()
        result_filter = self.result_filter.currentData()
        search_text = self.search_input.text().lower()

        for row in range(self.objects_table.rowCount()):
            should_show = True

            # Фильтр по подразделению
            if department_filter != "Все подразделения":
                dept_item = self.objects_table.item(row, 2)
                if dept_item.text() != department_filter:
                    should_show = False

            # Фильтр по результату
            if result_filter and should_show:
                result_item = self.objects_table.item(row, 6)
                if result_filter == "no_conclusion":
                    if result_item.text() != "Нет заключения":
                        should_show = False
                elif result_item.text() != result_filter:
                    should_show = False

            # Фильтр по поиску
            if should_show and search_text:
                name_item = self.objects_table.item(row, 0)
                number_item = self.objects_table.item(row, 1)
                
                if (
                    search_text not in name_item.text().lower()
                    and search_text not in number_item.text().lower()
                ):
                    should_show = False

            self.objects_table.setRowHidden(row, not should_show)

    def add_object(self):
        dialog = ObjectDialog(
            equipment_data=self.equipment_data,
            specialists_data=self.specialists_data
        )
        if dialog.exec_() == QDialog.Accepted:
            try:
                self.objects_data.append(dialog.object_data)

                with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.objects_data, f, indent=4, ensure_ascii=False)

                self.update_objects_table()
                QMessageBox.information(self, "Успех", "Объект успешно добавлен")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось добавить объект: {str(e)}"
                )

    def edit_object(self):
        selected_row = self.objects_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите объект для редактирования")
            return

        object_id = self.objects_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующий объект в массиве данных
        found_index = -1
        for i, obj in enumerate(self.objects_data):
            if obj.get("object_id") == object_id:
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(self, "Ошибка", "Выбранный объект не найден в базе данных")
            return

        object_data = self.objects_data[found_index]
        dialog = ObjectDialog(
            object_data=object_data,
            equipment_data=self.equipment_data,
            specialists_data=self.specialists_data
        )

        if dialog.exec_() == QDialog.Accepted:
            try:
                self.objects_data[found_index] = dialog.object_data

                with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.objects_data, f, indent=4, ensure_ascii=False)

                self.update_objects_table()
                QMessageBox.information(self, "Успех", "Изменения сохранены")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
                )

    def delete_object(self):
        if self.user_role not in ["admin", "manager"]:
            QMessageBox.warning(self, "Ошибка", "Недостаточно прав для удаления объектов")
            return
            
        selected_row = self.objects_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите объект для удаления")
            return

        object_id = self.objects_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующий объект в массиве данных
        found_index = -1
        for i, obj in enumerate(self.objects_data):
            if obj.get("object_id") == object_id:
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(self, "Ошибка", "Выбранный объект не найден в базе данных")
            return

        object_data = self.objects_data[found_index]

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить объект '{object_data.get('name', '')}' (№{object_data.get('number', '')})?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                del self.objects_data[found_index]

                with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.objects_data, f, indent=4, ensure_ascii=False)

                self.update_objects_table()
                QMessageBox.information(self, "Успех", "Объект удален")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить объект: {str(e)}"
                )

    def view_object_details(self):
        selected_row = self.objects_table.currentRow()
        if selected_row == -1:
            return

        object_id = self.objects_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующий объект в массиве данных
        found_object = None
        for obj in self.objects_data:
            if obj.get("object_id") == object_id:
                found_object = obj
                break

        if not found_object:
            return

        # Создаем диалог с подробной информацией об объекте
        details_dialog = QDialog(self)
        details_dialog.setWindowTitle(f"Информация об объекте: {found_object.get('name', '')}")
        details_dialog.setMinimumSize(700, 500)

        layout = QVBoxLayout(details_dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Вкладки
        tabs = QTabWidget()
        
        # Основная информация
        info_tab = QWidget()
        info_layout = QFormLayout(info_tab)
        
        # Заголовок
        title = QLabel(found_object.get("name", ""))
        title.setStyleSheet("font: bold 18pt 'Segoe UI';")
        layout.addWidget(title)
        
        # Основные поля
        fields = [
            ("ID объекта:", found_object.get("object_id", "")),
            ("Номер:", found_object.get("number", "")),
            ("Подразделение:", found_object.get("department", "")),
            ("Заказчик:", found_object.get("customer", "")),
            ("Адрес:", found_object.get("address", "")),
        ]
        
        for label_text, value in fields:
            label = QLabel(label_text)
            label.setStyleSheet("font-weight: bold;")
            value_label = QLabel(value)
            info_layout.addRow(label, value_label)
        
        # Дата
        if "date" in found_object:
            date_label = QLabel("Дата создания:")
            date_label.setStyleSheet("font-weight: bold;")
            date_value = QLabel(
                datetime.strptime(found_object["date"], "%Y-%m-%d").strftime("%d.%m.%Y")
            )
            info_layout.addRow(date_label, date_value)
        
        # Примечания
        notes_label = QLabel("Примечания:")
        notes_label.setStyleSheet("font-weight: bold;")
        notes_value = QTextEdit()
        notes_value.setPlainText(found_object.get("notes", ""))
        notes_value.setReadOnly(True)
        notes_value.setMaximumHeight(80)
        info_layout.addRow(notes_label, notes_value)
        
        tabs.addTab(info_tab, "Основная информация")
        
        # Вкладка "Оборудование"
        equipment_tab = QWidget()
        equipment_layout = QVBoxLayout(equipment_tab)
        
        equipment_table = QTableWidget()
        equipment_table.setColumnCount(5)
        equipment_table.setHorizontalHeaderLabels(
            ["Наименование", "Инв. номер", "Сер. номер", "Подразделение", "Статус поверки"]
        )
        equipment_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Добавляем оборудование, если есть
        if "equipment" in found_object and found_object["equipment"]:
            equipment_ids = found_object["equipment"]
            equipment_data = []
            
            for eq in self.equipment_data:
                if eq.get("equipment_id") in equipment_ids:
                    equipment_data.append(eq)
            
            equipment_table.setRowCount(len(equipment_data))
            
            for i, eq in enumerate(equipment_data):
                equipment_table.setItem(i, 0, QTableWidgetItem(eq.get("name", "")))
                equipment_table.setItem(i, 1, QTableWidgetItem(eq.get("inventory_number", "")))
                equipment_table.setItem(i, 2, QTableWidgetItem(eq.get("serial_number", "")))
                equipment_table.setItem(i, 3, QTableWidgetItem(eq.get("department", "")))
                
                # Статус поверки
                today = datetime.now().date()
                next_date = datetime.strptime(eq["next_verification_date"], "%Y-%m-%d").date()
                days_left = (next_date - today).days
                
                status_item = QTableWidgetItem()
                if days_left < 0:
                    status_text = "Просрочено"
                    status_item.setForeground(QColor("#e74c3c"))
                elif days_left <= 30:
                    status_text = f"Истекает через {days_left} дней"
                    status_item.setForeground(QColor("#f39c12"))
                else:
                    status_text = f"Активно ({days_left} дней)"
                    status_item.setForeground(QColor("#2ecc71"))
                    
                status_item.setText(status_text)
                equipment_table.setItem(i, 4, status_item)
        else:
            equipment_table.setRowCount(1)
            no_data = QTableWidgetItem("Нет привязанного оборудования")
            no_data.setTextAlignment(Qt.AlignCenter)
            equipment_table.setSpan(0, 0, 1, 5)
            equipment_table.setItem(0, 0, no_data)
            
        equipment_layout.addWidget(equipment_table)
        tabs.addTab(equipment_tab, "Оборудование")
        
        # Вкладка "Специалисты"
        specialists_tab = QWidget()
        specialists_layout = QVBoxLayout(specialists_tab)
        
        specialists_table = QTableWidget()
        specialists_table.setColumnCount(4)
        specialists_table.setHorizontalHeaderLabels(
            ["ФИО", "Должность", "Методы контроля", "Статус аттестации"]
        )
        specialists_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Добавляем специалистов, если есть
        if "specialists" in found_object and found_object["specialists"]:
            specialists_ids = found_object["specialists"]
            specialists_data = []
            
            for spec in self.specialists_data:
                if spec.get("specialist_id") in specialists_ids:
                    specialists_data.append(spec)
            
            specialists_table.setRowCount(len(specialists_data))
            
            for i, spec in enumerate(specialists_data):
                specialists_table.setItem(i, 0, QTableWidgetItem(spec.get("full_name", "")))
                specialists_table.setItem(i, 1, QTableWidgetItem(spec.get("position", "")))
                
                # Методы контроля
                methods = spec.get("methods", {})
                methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
                specialists_table.setItem(i, 2, QTableWidgetItem(methods_text))
                
                # Статус аттестации
                if "next_cert_date" in spec:
                    today = datetime.now().date()
                    next_date = datetime.strptime(spec["next_cert_date"], "%Y-%m-%d").date()
                    days_left = (next_date - today).days
                    
                    status_item = QTableWidgetItem()
                    if days_left < 0:
                        status_text = "Просрочено"
                        status_item.setForeground(QColor("#e74c3c"))
                    elif days_left <= 30:
                        status_text = f"Истекает через {days_left} дней"
                        status_item.setForeground(QColor("#f39c12"))
                    else:
                        status_text = f"Активно ({days_left} дней)"
                        status_item.setForeground(QColor("#2ecc71"))
                        
                    status_item.setText(status_text)
                    specialists_table.setItem(i, 3, status_item)
                else:
                    specialists_table.setItem(i, 3, QTableWidgetItem("Не указан"))
        else:
            specialists_table.setRowCount(1)
            no_data = QTableWidgetItem("Нет назначенных специалистов")
            no_data.setTextAlignment(Qt.AlignCenter)
            specialists_table.setSpan(0, 0, 1, 4)
            specialists_table.setItem(0, 0, no_data)
            
        specialists_layout.addWidget(specialists_table)
        tabs.addTab(specialists_tab, "Специалисты")
        
        # Вкладка "Заключение"
        conclusion_tab = QWidget()
        conclusion_layout = QVBoxLayout(conclusion_tab)
        
        conclusion = found_object.get("conclusion", {})
        if conclusion and conclusion.get("number") and conclusion.get("text"):
            # Информация о заключении
            conclusion_info = QFormLayout()
            
            conclusion_info.addRow("Номер заключения:", QLabel(conclusion.get("number", "")))
            
            if "date" in conclusion:
                conclusion_info.addRow(
                    "Дата заключения:", 
                    QLabel(datetime.strptime(conclusion["date"], "%Y-%m-%d").strftime("%d.%m.%Y"))
                )
                
            result_label = QLabel(conclusion.get("result", ""))
            if conclusion.get("result") == "Годен":
                result_label.setStyleSheet("color: #2ecc71; font-weight: bold;")
            elif conclusion.get("result") == "Не годен":
                result_label.setStyleSheet("color: #e74c3c; font-weight: bold;")
            elif conclusion.get("result") == "Требуется доработка":
                result_label.setStyleSheet("color: #f39c12; font-weight: bold;")
                
            conclusion_info.addRow("Результат:", result_label)
            
            conclusion_layout.addLayout(conclusion_info)
            
            # Текст заключения
            conclusion_layout.addWidget(QLabel("Текст заключения:"))
            conclusion_text = QTextEdit()
            conclusion_text.setPlainText(conclusion.get("text", ""))
            conclusion_text.setReadOnly(True)
            conclusion_layout.addWidget(conclusion_text)
            
            # Кнопка печати заключения
            print_btn = QPushButton("Печать заключения")
            print_btn.clicked.connect(lambda: self.print_conclusion(found_object))
            conclusion_layout.addWidget(print_btn)
        else:
            no_conclusion = QLabel("Заключение не сформировано")
            no_conclusion.setAlignment(Qt.AlignCenter)
            no_conclusion.setStyleSheet("font: italic 12pt 'Segoe UI';")
            conclusion_layout.addWidget(no_conclusion)
            
        tabs.addTab(conclusion_tab, "Заключение")
        
        layout.addWidget(tabs)
        
        # Кнопка закрытия
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(details_dialog.accept)
        layout.addWidget(close_btn)
        
        details_dialog.exec_()

    def generate_conclusion(self):
        selected_row = self.objects_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите объект для формирования заключения")
            return

        object_id = self.objects_table.item(selected_row, 0).data(Qt.UserRole)

        # Ищем соответствующий объект в массиве данных
        found_index = -1
        for i, obj in enumerate(self.objects_data):
            if obj.get("object_id") == object_id:
                found_index = i
                break

        if found_index == -1:
            QMessageBox.warning(self, "Ошибка", "Выбранный объект не найден в базе данных")
            return

        object_data = self.objects_data[found_index]
        
        # Открываем редактор заключения
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Формирование заключения: {object_data.get('name', '')}")
        dialog.setMinimumSize(700, 500)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Информация об объекте
        info_layout = QFormLayout()
        info_layout.addRow("Наименование объекта:", QLabel(object_data.get("name", "")))
        info_layout.addRow("Номер объекта:", QLabel(object_data.get("number", "")))
        info_layout.addRow("Заказчик:", QLabel(object_data.get("customer", "")))
        layout.addLayout(info_layout)
        
        # Данные заключения
        conclusion_layout = QFormLayout()
        
        conclusion_number = QLineEdit()
        conclusion_date = QDateEdit()
        conclusion_date.setCalendarPopup(True)
        conclusion_date.setDate(datetime.now().date())
        
        conclusion_result = QComboBox()
        conclusion_result.addItems(["Годен", "Не годен", "Требуется доработка"])
        
        # Если уже есть заключение, заполняем поля
        conclusion = object_data.get("conclusion", {})
        if conclusion:
            conclusion_number.setText(conclusion.get("number", ""))
            
            if "date" in conclusion:
                conclusion_date.setDate(
                    datetime.strptime(conclusion["date"], "%Y-%m-%d").date()
                )
                
            result_index = conclusion_result.findText(conclusion.get("result", ""))
            if result_index >= 0:
                conclusion_result.setCurrentIndex(result_index)
        
        conclusion_layout.addRow("Номер заключения:", conclusion_number)
        conclusion_layout.addRow("Дата заключения:", conclusion_date)
        conclusion_layout.addRow("Результат:", conclusion_result)
        
        layout.addLayout(conclusion_layout)
        
        # Текст заключения
        layout.addWidget(QLabel("Текст заключения:"))
        
        conclusion_text = QTextEdit()
        if conclusion and "text" in conclusion:
            conclusion_text.setPlainText(conclusion["text"])
        else:
            # Генерируем шаблон заключения
            template = self.generate_conclusion_template(object_data)
            conclusion_text.setPlainText(template)
            
        layout.addWidget(conclusion_text)
        
        # Кнопки
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("Сохранить")
        save_btn.clicked.connect(lambda: self.save_conclusion(
            dialog, 
            found_index, 
            conclusion_number.text(),
            conclusion_date.date(),
            conclusion_result.currentText(),
            conclusion_text.toPlainText()
        ))
        btn_layout.addWidget(save_btn)
        
        print_btn = QPushButton("Печать")
        print_btn.setEnabled(bool(conclusion_text.toPlainText()))
        print_btn.clicked.connect(lambda: self.print_text(conclusion_text.toPlainText()))
        btn_layout.addWidget(print_btn)
        
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec_()

    def generate_conclusion_template(self, object_data):
        # Собираем данные для шаблона
        specialists_data = []
        if "specialists" in object_data and object_data["specialists"]:
            for spec_id in object_data["specialists"]:
                for spec in self.specialists_data:
                    if spec.get("specialist_id") == spec_id:
                        specialists_data.append(spec)
                        break
        
        equipment_data = []
        if "equipment" in object_data and object_data["equipment"]:
            for eq_id in object_data["equipment"]:
                for eq in self.equipment_data:
                    if eq.get("equipment_id") == eq_id:
                        equipment_data.append(eq)
                        break
        
        # Формируем шаблон
        template = f"""ЗАКЛЮЧЕНИЕ № ____
по результатам неразрушающего контроля

1. Данные об объекте контроля:
   Наименование: {object_data.get('name', '')}
   Номер: {object_data.get('number', '')}
   Заказчик: {object_data.get('customer', '')}
   Адрес: {object_data.get('address', '')}
   Дата проведения контроля: {datetime.now().strftime('%d.%m.%Y')}

2. Методы контроля:
   {'- '.join([m for spec in specialists_data for m in spec.get('methods', {}).keys()])}

3. Оборудование:
"""

        for eq in equipment_data:
            template += f"   - {eq.get('name', '')} (инв. № {eq.get('inventory_number', '')})\n"
        
        template += """
4. Нормативная документация:
   - 

5. Результаты контроля:
   - 

6. Заключение:
   - 

7. Ответственные за проведение контроля:
"""

        for spec in specialists_data:
            template += f"   - {spec.get('full_name', '')}, {spec.get('position', '')}\n"
        
        template += f"""
Дата составления: {datetime.now().strftime('%d.%m.%Y')}

___________________
      (подпись)
"""
        return template

    def save_conclusion(self, dialog, object_index, number, date, result, text):
        if not number:
            QMessageBox.warning(dialog, "Ошибка", "Введите номер заключения")
            return
            
        if not text:
            QMessageBox.warning(dialog, "Ошибка", "Заполните текст заключения")
            return
        
        try:
            # Обновляем данные заключения в объекте
            conclusion = {
                "number": number,
                "date": date.toString("yyyy-MM-dd"),
                "result": result,
                "text": text
            }
            
            self.objects_data[object_index]["conclusion"] = conclusion
            
            # Сохраняем изменения в файл
            with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self.objects_data, f, indent=4, ensure_ascii=False)
            
            # Обновляем таблицу
            self.update_objects_table()
            
            QMessageBox.information(dialog, "Успех", "Заключение успешно сохранено")
            dialog.accept()
            
        except Exception as e:
            QMessageBox.critical(
                dialog, "Ошибка", f"Не удалось сохранить заключение: {str(e)}"
            )
            
    def print_text(self, text):
        if not text:
            QMessageBox.warning(self, "Предупреждение", "Нет текста для печати")
            return
            
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
        
        print_dialog = QPrintDialog(printer, self)
        if print_dialog.exec_() != QDialog.Accepted:
            return
            
        # Создаем документ для печати
        document = QTextEdit()
        document.setPlainText(text)
        document.print_(printer)
        
    def print_conclusion(self, object_data):
        conclusion = object_data.get("conclusion", {})
        if not conclusion or not conclusion.get("text"):
            QMessageBox.warning(self, "Предупреждение", "Заключение не сформировано")
            return
            
        self.print_text(conclusion.get("text", ""))
            
    def export_objects(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт данных",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
            options=options,
        )

        if not file_name:
            return

        try:
            if file_name.endswith(".xlsx"):
                self.export_to_excel(file_name)
            else:
                self.export_to_csv(file_name)

            QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось экспортировать данные: {str(e)}"
            )

    def export_to_excel(self, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Объекты"

        # Заголовки
        headers = [
            "Наименование",
            "Номер",
            "Подразделение",
            "Дата",
            "Заказчик",
            "Адрес",
            "Номер заключения",
            "Дата заключения",
            "Результат",
            "Примечания",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Данные
        for row, obj in enumerate(self.objects_data, 2):
            sheet.cell(row=row, column=1, value=obj.get("name", ""))
            sheet.cell(row=row, column=2, value=obj.get("number", ""))
            sheet.cell(row=row, column=3, value=obj.get("department", ""))
            
            if "date" in obj:
                date_obj = datetime.strptime(obj["date"], "%Y-%m-%d").date()
                sheet.cell(row=row, column=4, value=date_obj.strftime("%d.%m.%Y"))
                
            sheet.cell(row=row, column=5, value=obj.get("customer", ""))
            sheet.cell(row=row, column=6, value=obj.get("address", ""))
            
            conclusion = obj.get("conclusion", {})
            if conclusion:
                sheet.cell(row=row, column=7, value=conclusion.get("number", ""))
                
                if "date" in conclusion:
                    concl_date = datetime.strptime(conclusion["date"], "%Y-%m-%d").date()
                    sheet.cell(row=row, column=8, value=concl_date.strftime("%d.%m.%Y"))
                    
                sheet.cell(row=row, column=9, value=conclusion.get("result", ""))
                
            sheet.cell(row=row, column=10, value=obj.get("notes", ""))

        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Границы для всех ячеек
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        workbook.save(file_path)

    def export_to_csv(self, file_path):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            # Заголовки
            writer.writerow(
                [
                    "Наименование",
                    "Номер",
                    "Подразделение",
                    "Дата",
                    "Заказчик",
                    "Адрес",
                    "Номер заключения",
                    "Дата заключения",
                    "Результат",
                    "Примечания",
                ]
            )

            # Данные
            for obj in self.objects_data:
                conclusion = obj.get("conclusion", {})
                
                row_data = [
                    obj.get("name", ""),
                    obj.get("number", ""),
                    obj.get("department", ""),
                    datetime.strptime(obj["date"], "%Y-%m-%d").strftime("%d.%m.%Y") if "date" in obj else "",
                    obj.get("customer", ""),
                    obj.get("address", ""),
                    conclusion.get("number", ""),
                    datetime.strptime(conclusion["date"], "%Y-%m-%d").strftime("%d.%m.%Y") if "date" in conclusion else "",
                    conclusion.get("result", ""),
                    obj.get("notes", "")
                ]
                
                writer.writerow(row_data)


class EquipmentControlApp(QMainWindow):
    def __init__(self, user_role="guest", user_email="", username=""):
        super().__init__()
        self.user_role = user_role
        self.user_email = user_email
        self.username = username
        self.equipment_data = []
        self.notification_timer = QTimer(self)
        self.notification_timer.timeout.connect(self.check_verification_dates)
        self.notification_timer.start(86400000)  # Проверка раз в день (24 часа)

        try:
            self.init_ui()
            self.load_data()
            self.check_verification_dates()
        except Exception as e:
            QMessageBox.critical(
                None,
                "Ошибка инициализации",
                f"Не удалось инициализировать приложение: {str(e)}",
            )
            sys.exit(1)

    def init_ui(self):
        self.setWindowTitle("ЛНК. ООО ЮНИКС. Контроль поверок оборудования. Разработчик Тешуков А.В.")
        self.setMinimumSize(1200, 800)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Основной layout
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Боковая панель
        self.setup_sidebar(main_layout)

        # Область контента
        self.content_area = QStackedWidget()
        main_layout.addWidget(self.content_area)

        # Страницы приложения
        self.setup_pages()

        # Применение стилей
        self.apply_styles()

    def setup_sidebar(self, main_layout):
        sidebar = QFrame()
        sidebar.setFixedWidth(220)
        sidebar.setStyleSheet("""
            QFrame {
                background-color: #2c3e50;
                border: none;
            }
            QLabel {
                color: white;
                font: bold 16pt 'Segoe UI';
            }
            QPushButton {
                background-color: #34495e;
                color: white;
                border: none;
                border-radius: 5px;
                min-height: 40px;
                font: 12pt 'Segoe UI';
                padding: 5px 15px;
            }
            QPushButton:hover {
                background-color: #4a90e2;
            }
        """)

        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(10, 20, 10, 20)
        sidebar_layout.setSpacing(15)

        # Логотип и информация о пользователе
        logo = QLabel("Поверки")
        logo.setAlignment(Qt.AlignCenter)
        logo.setFixedHeight(60)
        sidebar_layout.addWidget(logo)

        user_info = QLabel(
            f"Пользователь: {self.username}\nРоль: {ROLES.get(self.user_role, 'Гость')}"
        )
        user_info.setAlignment(Qt.AlignCenter)
        user_info.setWordWrap(True)
        sidebar_layout.addWidget(user_info)

        # Кнопки навигации
        nav_buttons = [
            ("Главная", self.show_dashboard),
            ("Оборудование", self.show_equipment),
            ("Специалисты", self.show_specialists),
            ("Объекты", self.show_objects),
            ("Отчеты", self.show_reports),
        ]

        for text, handler in nav_buttons:
            btn = QPushButton(text)
            btn.clicked.connect(handler)
            sidebar_layout.addWidget(btn)

        # Кнопка управления пользователями (только для администратора)
        if self.user_role == "admin":
            self.user_management_btn = QPushButton("Упр-ние админом")
            self.user_management_btn.clicked.connect(self.manage_users)
            sidebar_layout.addWidget(self.user_management_btn)

        # Кнопка ИИ-помощника
        self.ai_btn = QPushButton("ИИ-помощник")
        self.ai_btn.clicked.connect(self.show_ai_chat)
        sidebar_layout.addWidget(self.ai_btn)

        sidebar_layout.addStretch()

        # Кнопка выхода
        btn_logout = QPushButton("Выход")
        btn_logout.setStyleSheet("background-color: #e74c3c;")
        btn_logout.clicked.connect(self.close_application)
        sidebar_layout.addWidget(btn_logout)

        main_layout.addWidget(sidebar)

    def setup_pages(self):
        # Главная страница
        self.dashboard_page = self.create_dashboard_page()
        self.content_area.addWidget(self.dashboard_page)

        # Страница оборудования
        self.equipment_page = self.create_equipment_page()
        self.content_area.addWidget(self.equipment_page)
        
        # Страница специалистов
        self.specialists_page = SpecialistsWidget(user_role=self.user_role)
        self.content_area.addWidget(self.specialists_page)
        
        # Страница объектов
        self.objects_page = ObjectsWidget(user_role=self.user_role)
        self.content_area.addWidget(self.objects_page)

        # Страница отчетов
        self.reports_page = self.create_reports_page()
        self.content_area.addWidget(self.reports_page)

        # Страница ИИ-помощника
        self.ai_chat_page = AIChatWidget()
        self.content_area.addWidget(self.ai_chat_page)

    def create_dashboard_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Главная")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        layout.addWidget(title)

        # Статистика
        stats_layout = QGridLayout()
        stats_layout.setSpacing(15)

        # Карточка с количеством оборудования
        equipment_card = QFrame()
        equipment_card.setFixedHeight(120)
        equipment_card.setStyleSheet("""
            QFrame {
                background-color: #3498db;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        equipment_layout = QVBoxLayout(equipment_card)
        equipment_layout.setContentsMargins(15, 15, 15, 15)
        self.equipment_count_label = QLabel("0")
        self.equipment_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.equipment_count_label.setAlignment(Qt.AlignCenter)
        equipment_layout.addWidget(QLabel("Всего оборудования"))
        equipment_layout.addWidget(self.equipment_count_label)
        stats_layout.addWidget(equipment_card, 0, 0)

        # Карточка с истекающими поверками
        expiring_card = QFrame()
        expiring_card.setFixedHeight(120)
        expiring_card.setStyleSheet("""
            QFrame {
                background-color: #e74c3c;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        expiring_layout = QVBoxLayout(expiring_card)
        expiring_layout.setContentsMargins(15, 15, 15, 15)
        self.expiring_count_label = QLabel("0")
        self.expiring_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.expiring_count_label.setAlignment(Qt.AlignCenter)
        expiring_layout.addWidget(QLabel("Истекающие поверки"))
        expiring_layout.addWidget(self.expiring_count_label)
        stats_layout.addWidget(expiring_card, 0, 1)

        # Карточка с просроченными поверками
        overdue_card = QFrame()
        overdue_card.setFixedHeight(120)
        overdue_card.setStyleSheet("""
            QFrame {
                background-color: #f39c12;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        overdue_layout = QVBoxLayout(overdue_card)
        overdue_layout.setContentsMargins(15, 15, 15, 15)
        self.overdue_count_label = QLabel("0")
        self.overdue_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.overdue_count_label.setAlignment(Qt.AlignCenter)
        overdue_layout.addWidget(QLabel("Просроченные поверки"))
        overdue_layout.addWidget(self.overdue_count_label)
        stats_layout.addWidget(overdue_card, 0, 2)
        
        # Карточка с количеством специалистов
        specialists_card = QFrame()
        specialists_card.setFixedHeight(120)
        specialists_card.setStyleSheet("""
            QFrame {
                background-color: #2ecc71;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        specialists_layout = QVBoxLayout(specialists_card)
        specialists_layout.setContentsMargins(15, 15, 15, 15)
        self.specialists_count_label = QLabel("0")
        self.specialists_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.specialists_count_label.setAlignment(Qt.AlignCenter)
        specialists_layout.addWidget(QLabel("Всего специалистов"))
        specialists_layout.addWidget(self.specialists_count_label)
        stats_layout.addWidget(specialists_card, 1, 0)
        
        # Карточка с истекающими удостоверениями
        expiring_cert_card = QFrame()
        expiring_cert_card.setFixedHeight(120)
        expiring_cert_card.setStyleSheet("""
            QFrame {
                background-color: #9b59b6;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        expiring_cert_layout = QVBoxLayout(expiring_cert_card)
        expiring_cert_layout.setContentsMargins(15, 15, 15, 15)
        self.expiring_cert_count_label = QLabel("0")
        self.expiring_cert_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.expiring_cert_count_label.setAlignment(Qt.AlignCenter)
        expiring_cert_layout.addWidget(QLabel("Истекающие удостоверения"))
        expiring_cert_layout.addWidget(self.expiring_cert_count_label)
        stats_layout.addWidget(expiring_cert_card, 1, 1)
        
        # Карточка с количеством объектов
        objects_card = QFrame()
        objects_card.setFixedHeight(120)
        objects_card.setStyleSheet("""
            QFrame {
                background-color: #1abc9c;
                border-radius: 10px;
            }
            QLabel {
                color: white;
                font: bold 14pt 'Segoe UI';
            }
        """)
        objects_layout = QVBoxLayout(objects_card)
        objects_layout.setContentsMargins(15, 15, 15, 15)
        self.objects_count_label = QLabel("0")
        self.objects_count_label.setStyleSheet("font: bold 24pt 'Segoe UI';")
        self.objects_count_label.setAlignment(Qt.AlignCenter)
        objects_layout.addWidget(QLabel("Всего объектов"))
        objects_layout.addWidget(self.objects_count_label)
        stats_layout.addWidget(objects_card, 1, 2)

        layout.addLayout(stats_layout)

        # Таблица с ближайшими поверками
        self.verification_table = QTableWidget()
        self.verification_table.setColumnCount(5)
        self.verification_table.setHorizontalHeaderLabels(
            ["Наименование", "Инв. номер", "Подразделение", "Дата поверки", "Статус"]
        )
        self.verification_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.verification_table.verticalHeader().setVisible(False)
        self.verification_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(QLabel("Ближайшие поверки:"))
        layout.addWidget(self.verification_table)
        
        # Таблица с истекающими удостоверениями
        self.cert_table = QTableWidget()
        self.cert_table.setColumnCount(5)
        self.cert_table.setHorizontalHeaderLabels(
            ["ФИО", "Должность", "Подразделение", "Срок действия до", "Статус"]
        )
        self.cert_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.cert_table.verticalHeader().setVisible(False)
        self.cert_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(QLabel("Ближайшие истечения удостоверений:"))
        layout.addWidget(self.cert_table)

        return page

    def create_equipment_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок и кнопки
        header_layout = QHBoxLayout()
        title = QLabel("Оборудование")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        header_layout.addWidget(title)

        # Кнопки управления (только для администратора и метролога)
        if self.user_role in ["admin", "metrolog"]:
            btn_layout = QHBoxLayout()
            btn_layout.setSpacing(10)

            add_btn = QPushButton("Добавить")
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            add_btn.clicked.connect(self.add_equipment)
            btn_layout.addWidget(add_btn)

            edit_btn = QPushButton("Редактировать")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #0b7dda;
                }
            """)
            edit_btn.clicked.connect(self.edit_equipment)
            btn_layout.addWidget(edit_btn)

            delete_btn = QPushButton("Удалить")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 15px;
                    font: bold 12pt 'Segoe UI';
                }
                QPushButton:hover {
                    background-color: #d32f2f;
                }
            """)
            delete_btn.clicked.connect(self.delete_equipment)
            btn_layout.addWidget(delete_btn)

            header_layout.addLayout(btn_layout)

        layout.addLayout(header_layout)

        # Фильтры
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.department_filter = QComboBox()
        self.department_filter.addItem("Все подразделения")
        for dept in DEPARTMENTS:
            self.department_filter.addItem(dept)
        self.department_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Подразделение:"))
        filter_layout.addWidget(self.department_filter)

        self.status_filter = QComboBox()
        self.status_filter.addItem("Все статусы", "")
        self.status_filter.addItem("Активные", "active")
        self.status_filter.addItem("Истекающие", "expiring")
        self.status_filter.addItem("Просроченные", "overdue")
        self.status_filter.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Статус:"))
        filter_layout.addWidget(self.status_filter)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по наименованию или номеру...")
        self.search_input.textChanged.connect(self.apply_filters)
        search_layout.addWidget(self.search_input)

        export_btn = QPushButton("Экспорт")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        export_btn.clicked.connect(self.export_equipment)
        search_layout.addWidget(export_btn)

        filter_layout.addLayout(search_layout)
        layout.addLayout(filter_layout)

        # Таблица оборудования
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(7)
        self.equipment_table.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Инв. номер",
                "Сер. номер",
                "Подразделение",
                "Дата поверки",
                "След. поверка",
                "Статус",
            ]
        )
        self.equipment_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )
        self.equipment_table.verticalHeader().setVisible(False)
        self.equipment_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.equipment_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.equipment_table.setSelectionMode(QTableWidget.SingleSelection)
        self.equipment_table.setSortingEnabled(True)
        self.equipment_table.doubleClicked.connect(self.view_equipment_details)
        layout.addWidget(self.equipment_table)

        return page

    def create_reports_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Заголовок
        title = QLabel("Отчеты")
        title.setStyleSheet("font: bold 24pt 'Segoe UI';")
        layout.addWidget(title)

        # Вкладки для типов отчетов
        report_tabs = QTabWidget()
        
        # Вкладка отчетов по оборудованию
        equipment_report_tab = QWidget()
        eq_report_layout = QVBoxLayout(equipment_report_tab)

        # Кнопки отчетов по оборудованию
        eq_reports_layout = QHBoxLayout()
        eq_reports_layout.setSpacing(15)

        # Кнопка для отчета по истекающим поверкам
        expiring_btn = QPushButton("Истекающие поверки")
        expiring_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        expiring_btn.clicked.connect(lambda: self.generate_equipment_report("expiring"))
        eq_reports_layout.addWidget(expiring_btn)

        # Кнопка для отчета по просроченным поверкам
        overdue_btn = QPushButton("Просроченные поверки")
        overdue_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        overdue_btn.clicked.connect(lambda: self.generate_equipment_report("overdue"))
        eq_reports_layout.addWidget(overdue_btn)

        # Кнопка для полного отчета
        full_report_btn = QPushButton("Полный отчет")
        full_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        full_report_btn.clicked.connect(lambda: self.generate_equipment_report("full"))
        eq_reports_layout.addWidget(full_report_btn)

        eq_report_layout.addLayout(eq_reports_layout)

        # Параметры отчета по оборудованию
        eq_params_layout = QFormLayout()
        eq_params_layout.setSpacing(15)

        # Выбор подразделения
        self.report_department_combo = QComboBox()
        self.report_department_combo.addItem("Все подразделения")
        self.report_department_combo.addItems(DEPARTMENTS)
        eq_params_layout.addRow("Подразделение:", self.report_department_combo)

        # Выбор формата экспорта
        self.export_format_combo = QComboBox()
        self.export_format_combo.addItem("Excel", "xlsx")
        self.export_format_combo.addItem("CSV", "csv")
        self.export_format_combo.addItem("PDF", "pdf")
        eq_params_layout.addRow("Формат:", self.export_format_combo)

        eq_report_layout.addLayout(eq_params_layout)

        # Область предпросмотра отчета по оборудованию
        self.equipment_report_preview = QTableWidget()
        self.equipment_report_preview.setColumnCount(6)
        self.equipment_report_preview.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Инв. номер",
                "Подразделение",
                "Дата поверки",
                "След. поверка",
                "Статус",
            ]
        )
        self.equipment_report_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.equipment_report_preview.verticalHeader().setVisible(False)
        self.equipment_report_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        eq_report_layout.addWidget(QLabel("Предпросмотр:"))
        eq_report_layout.addWidget(self.equipment_report_preview)
        
        report_tabs.addTab(equipment_report_tab, "Отчеты по оборудованию")
        
        # Вкладка отчетов по специалистам
        specialists_report_tab = QWidget()
        spec_report_layout = QVBoxLayout(specialists_report_tab)
        
        # Кнопки отчетов по специалистам
        spec_reports_layout = QHBoxLayout()
        spec_reports_layout.setSpacing(15)
        
        # Кнопка для отчета по истекающим удостоверениям
        expiring_cert_btn = QPushButton("Истекающие удостоверения")
        expiring_cert_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        expiring_cert_btn.clicked.connect(lambda: self.generate_specialists_report("expiring"))
        spec_reports_layout.addWidget(expiring_cert_btn)
        
        # Кнопка для отчета по методам контроля
        methods_btn = QPushButton("По методам контроля")
        methods_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        methods_btn.clicked.connect(lambda: self.generate_specialists_report("methods"))
        spec_reports_layout.addWidget(methods_btn)
        
        # Кнопка для полного отчета по специалистам
        full_spec_btn = QPushButton("Полный отчет")
        full_spec_btn.setStyleSheet("""
            QPushButton {
                background-color: #1abc9c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #16a085;
            }
        """)
        full_spec_btn.clicked.connect(lambda: self.generate_specialists_report("full"))
        spec_reports_layout.addWidget(full_spec_btn)
        
        spec_report_layout.addLayout(spec_reports_layout)
        
        # Параметры отчета по специалистам
        spec_params_layout = QFormLayout()
        spec_params_layout.setSpacing(15)
        
        # Выбор подразделения
        self.spec_report_department_combo = QComboBox()
        self.spec_report_department_combo.addItem("Все подразделения")
        self.spec_report_department_combo.addItems(DEPARTMENTS)
        spec_params_layout.addRow("Подразделение:", self.spec_report_department_combo)
        
        # Выбор метода контроля
        self.spec_report_method_combo = QComboBox()
        self.spec_report_method_combo.addItem("Все методы")
        self.spec_report_method_combo.addItems(CONTROL_METHODS)
        spec_params_layout.addRow("Метод контроля:", self.spec_report_method_combo)
        
        # Выбор формата экспорта
        self.spec_export_format_combo = QComboBox()
        self.spec_export_format_combo.addItem("Excel", "xlsx")
        self.spec_export_format_combo.addItem("CSV", "csv")
        self.spec_export_format_combo.addItem("PDF", "pdf")
        spec_params_layout.addRow("Формат:", self.spec_export_format_combo)
        
        spec_report_layout.addLayout(spec_params_layout)
        
        # Область предпросмотра отчета по специалистам
        self.specialists_report_preview = QTableWidget()
        self.specialists_report_preview.setColumnCount(6)
        self.specialists_report_preview.setHorizontalHeaderLabels(
            [
                "ФИО",
                "Должность",
                "Подразделение",
                "Методы контроля",
                "Срок действия",
                "Статус",
            ]
        )
        self.specialists_report_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.specialists_report_preview.verticalHeader().setVisible(False)
        self.specialists_report_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        spec_report_layout.addWidget(QLabel("Предпросмотр:"))
        spec_report_layout.addWidget(self.specialists_report_preview)
        
        report_tabs.addTab(specialists_report_tab, "Отчеты по специалистам")
        
        # Вкладка отчетов по объектам
        objects_report_tab = QWidget()
        obj_report_layout = QVBoxLayout(objects_report_tab)
        
        # Кнопки отчетов по объектам
        obj_reports_layout = QHBoxLayout()
        obj_reports_layout.setSpacing(15)
        
        # Кнопка для отчета по результатам
        results_btn = QPushButton("По результатам")
        results_btn.setStyleSheet("""
            QPushButton {
                background-color: #34495e;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #2c3e50;
            }
        """)
        results_btn.clicked.connect(lambda: self.generate_objects_report("results"))
        obj_reports_layout.addWidget(results_btn)
        
        # Кнопка для отчета по заказчикам
        customers_btn = QPushButton("По заказчикам")
        customers_btn.setStyleSheet("""
            QPushButton {
                background-color: #f1c40f;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #f39c12;
            }
        """)
        customers_btn.clicked.connect(lambda: self.generate_objects_report("customers"))
        obj_reports_layout.addWidget(customers_btn)
        
        # Кнопка для полного отчета по объектам
        full_obj_btn = QPushButton("Полный отчет")
        full_obj_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 15px;
                font: bold 12pt 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        full_obj_btn.clicked.connect(lambda: self.generate_objects_report("full"))
        obj_reports_layout.addWidget(full_obj_btn)
        
        obj_report_layout.addLayout(obj_reports_layout)
        
        # Параметры отчета по объектам
        obj_params_layout = QFormLayout()
        obj_params_layout.setSpacing(15)
        
        # Выбор подразделения
        self.obj_report_department_combo = QComboBox()
        self.obj_report_department_combo.addItem("Все подразделения")
        self.obj_report_department_combo.addItems(DEPARTMENTS)
        obj_params_layout.addRow("Подразделение:", self.obj_report_department_combo)
        
        # Выбор результата
        self.obj_report_result_combo = QComboBox()
        self.obj_report_result_combo.addItem("Все результаты")
        self.obj_report_result_combo.addItems(["Годен", "Не годен", "Требуется доработка", "Нет заключения"])
        obj_params_layout.addRow("Результат:", self.obj_report_result_combo)
        
        # Диапазон дат
        date_range_layout = QHBoxLayout()
        
        self.obj_report_start_date = QDateEdit()
        self.obj_report_start_date.setCalendarPopup(True)
        self.obj_report_start_date.setDate(datetime.now().date().replace(month=1, day=1))
        
        self.obj_report_end_date = QDateEdit()
        self.obj_report_end_date.setCalendarPopup(True)
        self.obj_report_end_date.setDate(datetime.now().date())
        
        date_range_layout.addWidget(QLabel("С:"))
        date_range_layout.addWidget(self.obj_report_start_date)
        date_range_layout.addWidget(QLabel("По:"))
        date_range_layout.addWidget(self.obj_report_end_date)
        
        obj_params_layout.addRow("Период:", date_range_layout)
        
        # Выбор формата экспорта
        self.obj_export_format_combo = QComboBox()
        self.obj_export_format_combo.addItem("Excel", "xlsx")
        self.obj_export_format_combo.addItem("CSV", "csv")
        self.obj_export_format_combo.addItem("PDF", "pdf")
        obj_params_layout.addRow("Формат:", self.obj_export_format_combo)
        
        obj_report_layout.addLayout(obj_params_layout)
        
        # Область предпросмотра отчета по объектам
        self.objects_report_preview = QTableWidget()
        self.objects_report_preview.setColumnCount(6)
        self.objects_report_preview.setHorizontalHeaderLabels(
            [
                "Наименование",
                "Номер",
                "Дата",
                "Заказчик",
                "Номер заключения",
                "Результат",
            ]
        )
        self.objects_report_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.objects_report_preview.verticalHeader().setVisible(False)
        self.objects_report_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        obj_report_layout.addWidget(QLabel("Предпросмотр:"))
        obj_report_layout.addWidget(self.objects_report_preview)
        
        report_tabs.addTab(objects_report_tab, "Отчеты по объектам")
        
        layout.addWidget(report_tabs)

        return page

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                font: 11pt 'Segoe UI';
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 5px;
                font: bold 11pt 'Segoe UI';
            }
            QLineEdit, QComboBox {
                padding: 5px;
                font: 11pt 'Segoe UI';
                border: 1px solid #ddd;
                border-radius: 3px;
            }
            QLabel {
                font: 11pt 'Segoe UI';
            }
        """)

    def load_data(self):
        try:
            # Загружаем данные оборудования
            if not os.path.exists(EQUIPMENT_DATA_FILE):
                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump([], f)

            with open(EQUIPMENT_DATA_FILE, "r", encoding="utf-8") as f:
                self.equipment_data = json.load(f)
            
            # Загружаем данные специалистов
            if os.path.exists(SPECIALISTS_DATA_FILE):
                with open(SPECIALISTS_DATA_FILE, "r", encoding="utf-8") as f:
                    self.specialists_data = json.load(f)
            else:
                self.specialists_data = []
                with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.specialists_data, f)
            
            # Загружаем данные объектов
            if os.path.exists(OBJECTS_DATA_FILE):
                with open(OBJECTS_DATA_FILE, "r", encoding="utf-8") as f:
                    self.objects_data = json.load(f)
            else:
                self.objects_data = []
                with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.objects_data, f)

            # Обновляем таблицу оборудования и информацию на главной странице
            self.update_equipment_table()
            self.update_dashboard()
            
            # Обновляем страницу специалистов и объектов
            self.specialists_page.load_data()
            self.objects_page.load_data()

        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка загрузки",
                f"Не удалось загрузить данные: {str(e)}",
            )

    def update_equipment_table(self):
        self.equipment_table.setRowCount(len(self.equipment_data))

        for row, equipment in enumerate(self.equipment_data):
            # Наименование
            name_item = QTableWidgetItem(equipment["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            name_item.setData(Qt.UserRole, equipment.get("equipment_id", ""))
            self.equipment_table.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(equipment["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 1, inv_item)

            # Серийный номер
            serial_item = QTableWidgetItem(equipment["serial_number"])
            serial_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 2, serial_item)

            # Подразделение
            dept_item = QTableWidgetItem(equipment["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 3, dept_item)

            # Дата поверки
            verif_date = datetime.strptime(
                equipment["verification_date"], "%Y-%m-%d"
            ).date()
            verif_item = QTableWidgetItem(verif_date.strftime("%d.%m.%Y"))
            verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 4, verif_item)

            # Следующая поверка
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            next_verif_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            next_verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_table.setItem(row, 5, next_verif_item)

            # Статус
            status, color = self.get_verification_status(next_verif_date)
            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(color))
            self.equipment_table.setItem(row, 6, status_item)

    def update_dashboard(self):
        # Общее количество оборудования
        self.equipment_count_label.setText(str(len(self.equipment_data)))

        # Количество истекающих и просроченных поверок
        expiring_count = 0
        overdue_count = 0
        today = datetime.now().date()

        verification_data = []

        for equipment in self.equipment_data:
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if days_left < 0:
                overdue_count += 1
                status = "Просрочено"
                color = "#e74c3c"
            elif days_left <= 30:
                expiring_count += 1
                status = f"Истекает через {days_left} дней"
                color = "#f39c12"
            else:
                status = f"Активно ({days_left} дней)"
                color = "#2ecc71"

            verification_data.append(
                {
                    "equipment": equipment,
                    "days_left": days_left,
                    "status": status,
                    "color": color,
                }
            )

        self.expiring_count_label.setText(str(expiring_count))
        self.overdue_count_label.setText(str(overdue_count))

        # Данные специалистов
        specialists_count = len(self.specialists_data)
        self.specialists_count_label.setText(str(specialists_count))
        
        # Количество истекающих удостоверений
        expiring_cert_count = 0
        cert_data = []
        
        for specialist in self.specialists_data:
            if "next_cert_date" in specialist:
                next_cert_date = datetime.strptime(
                    specialist["next_cert_date"], "%Y-%m-%d"
                ).date()
                days_left = (next_cert_date - today).days
                
                if days_left < 0:
                    status = "Просрочено"
                    color = "#e74c3c"
                elif days_left <= 30:
                    expiring_cert_count += 1
                    status = f"Истекает через {days_left} дней"
                    color = "#f39c12"
                else:
                    status = f"Активно ({days_left} дней)"
                    color = "#2ecc71"
                    
                cert_data.append({
                    "specialist": specialist,
                    "days_left": days_left,
                    "status": status,
                    "color": color
                })
        
        self.expiring_cert_count_label.setText(str(expiring_cert_count))
        
        # Количество объектов
        objects_count = len(self.objects_data)
        self.objects_count_label.setText(str(objects_count))

        # Сортируем по дням до поверки (ближайшие сверху)
        verification_data.sort(key=lambda x: x["days_left"])
        cert_data.sort(key=lambda x: x["days_left"])

        # Обновляем таблицу ближайших поверок
        self.verification_table.setRowCount(min(10, len(verification_data)))

        for row, data in enumerate(verification_data[:10]):
            eq = data["equipment"]

            # Наименование
            name_item = QTableWidgetItem(eq["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.verification_table.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(eq["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 1, inv_item)

            # Подразделение
            dept_item = QTableWidgetItem(eq["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 2, dept_item)

            # Дата следующей поверки
            next_verif_date = datetime.strptime(
                eq["next_verification_date"], "%Y-%m-%d"
            ).date()
            date_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            date_item.setTextAlignment(Qt.AlignCenter)
            self.verification_table.setItem(row, 3, date_item)

            # Статус
            status_item = QTableWidgetItem(data["status"])
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(data["color"]))
            self.verification_table.setItem(row, 4, status_item)
            
        # Обновляем таблицу истекающих удостоверений
        self.cert_table.setRowCount(min(10, len(cert_data)))
        
        for row, data in enumerate(cert_data[:10]):
            spec = data["specialist"]
            
            # ФИО
            name_item = QTableWidgetItem(spec.get("full_name", ""))
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.cert_table.setItem(row, 0, name_item)
            
            # Должность
            position_item = QTableWidgetItem(spec.get("position", ""))
            position_item.setTextAlignment(Qt.AlignCenter)
            self.cert_table.setItem(row, 1, position_item)
            
            # Подразделение
            dept_item = QTableWidgetItem(spec.get("department", ""))
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.cert_table.setItem(row, 2, dept_item)
            
            # Срок действия
            next_cert_date = datetime.strptime(
                spec["next_cert_date"], "%Y-%m-%d"
            ).date()
            date_item = QTableWidgetItem(next_cert_date.strftime("%d.%m.%Y"))
            date_item.setTextAlignment(Qt.AlignCenter)
            self.cert_table.setItem(row, 3, date_item)
            
            # Статус
            status_item = QTableWidgetItem(data["status"])
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(data["color"]))
            self.cert_table.setItem(row, 4, status_item)

    def get_verification_status(self, next_verif_date):
        today = datetime.now().date()
        days_left = (next_verif_date - today).days

        if days_left < 0:
            return "Просрочено", "#e74c3c"
        elif days_left <= 30:
            return f"Истекает ({days_left} дн.)", "#f39c12"
        else:
            return f"Активно ({days_left} дн.)", "#2ecc71"

    def apply_filters(self):
        department_filter = self.department_filter.currentText()
        status_filter = self.status_filter.currentData()
        search_text = self.search_input.text().lower()
        today = datetime.now().date()

        for row in range(self.equipment_table.rowCount()):
            should_show = True

            # Фильтр по подразделению
            if department_filter != "Все подразделения":
                dept_item = self.equipment_table.item(row, 3)
                if dept_item.text() != department_filter:
                    should_show = False

            # Фильтр по статусу
            if status_filter and should_show:
                next_verif_item = self.equipment_table.item(row, 5)
                next_verif_date = datetime.strptime(
                    next_verif_item.text(), "%d.%m.%Y"
                ).date()
                days_left = (next_verif_date - today).days

                if status_filter == "active" and days_left <= 30:
                    should_show = False
                elif status_filter == "expiring" and not (0 <= days_left <= 30):
                    should_show = False
                elif status_filter == "overdue" and days_left >= 0:
                    should_show = False

            # Фильтр по поиску
            if should_show and search_text:
                name_item = self.equipment_table.item(row, 0)
                inv_item = self.equipment_table.item(row, 1)
                serial_item = self.equipment_table.item(row, 2)

                if (
                    search_text not in name_item.text().lower()
                    and search_text not in inv_item.text().lower()
                    and search_text not in serial_item.text().lower()
                ):
                    should_show = False

            self.equipment_table.setRowHidden(row, not should_show)

    def view_equipment_details(self):
        selected_row = self.equipment_table.currentRow()
        if selected_row == -1:
            return
            
        equipment_id = self.equipment_table.item(selected_row, 0).data(Qt.UserRole)
        
        # Находим оборудование по ID
        found_equipment = None
        for eq in self.equipment_data:
            if eq.get("equipment_id") == equipment_id:
                found_equipment = eq
                break
                
        if not found_equipment:
            return
            
        # Показываем диалог с подробной информацией
        details_dialog = EquipmentDetailsDialog(found_equipment, self)
        details_dialog.exec_()

    def add_equipment(self):
        dialog = EquipmentDialog()
        if dialog.exec_() == QDialog.Accepted:
            try:
                self.equipment_data.append(dialog.equipment)

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Оборудование успешно добавлено")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось добавить оборудование: {str(e)}"
                )

    def edit_equipment(self):
        selected_row = self.equipment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выберите оборудование для редактирования"
            )
            return
            
        equipment_id = self.equipment_table.item(selected_row, 0).data(Qt.UserRole)
        
        # Находим оборудование по ID
        found_index = -1
        for i, eq in enumerate(self.equipment_data):
            if eq.get("equipment_id") == equipment_id:
                found_index = i
                break
                
        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранное оборудование не найдено в базе данных"
            )
            return

        equipment = self.equipment_data[found_index]
        dialog = EquipmentDialog(equipment, self)

        if dialog.exec_() == QDialog.Accepted:
            try:
                # Если изменился инвентарный номер, то нужно перенести документы
                if equipment["inventory_number"] != dialog.equipment["inventory_number"]:
                    old_doc_dir = os.path.join(DOCS_FOLDER, equipment["inventory_number"])
                    new_doc_dir = os.path.join(DOCS_FOLDER, dialog.equipment["inventory_number"])
                    
                    if os.path.exists(old_doc_dir):
                        # Создаем новую директорию, если её нет
                        os.makedirs(new_doc_dir, exist_ok=True)
                        
                        # Перемещаем все файлы
                        for doc in dialog.equipment.get("documents", []):
                            old_path = os.path.join(DOCS_FOLDER, doc["path"])
                            new_path = os.path.join(DOCS_FOLDER, dialog.equipment["inventory_number"], os.path.basename(doc["path"]))
                            
                            if os.path.exists(old_path):
                                shutil.copy2(old_path, new_path)
                                
                            # Обновляем путь в документе
                            doc["path"] = os.path.join(dialog.equipment["inventory_number"], os.path.basename(doc["path"]))
                
                # Если у редактируемого оборудования уже была поверка, добавляем её в историю
                if "verification_history" not in dialog.equipment:
                    dialog.equipment["verification_history"] = []
                    
                # Если текущая дата поверки отличается от сохраненной, добавляем старую в историю
                if equipment["verification_date"] != dialog.equipment["verification_date"]:
                    history_entry = {
                        "date": equipment["verification_date"],
                        "valid_until": equipment["next_verification_date"],
                        "notes": "Предыдущая поверка"
                    }
                    dialog.equipment["verification_history"].insert(0, history_entry)
                
                # Обновляем данные оборудования
                self.equipment_data[found_index] = dialog.equipment

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Изменения сохранены")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось сохранить изменения: {str(e)}"
                )

    def delete_equipment(self):
        selected_row = self.equipment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, "Ошибка", "Выберите оборудование для удаления")
            return
            
        equipment_id = self.equipment_table.item(selected_row, 0).data(Qt.UserRole)
        
        # Находим оборудование по ID
        found_index = -1
        for i, eq in enumerate(self.equipment_data):
            if eq.get("equipment_id") == equipment_id:
                found_index = i
                break
                
        if found_index == -1:
            QMessageBox.warning(
                self, "Ошибка", "Выбранное оборудование не найдено в базе данных"
            )
            return

        equipment = self.equipment_data[found_index]

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Вы уверены, что хотите удалить оборудование {equipment['name']} (инв. № {equipment['inventory_number']})?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            try:
                # Удаляем директорию с документами
                doc_dir = os.path.join(DOCS_FOLDER, equipment["inventory_number"])
                if os.path.exists(doc_dir):
                    shutil.rmtree(doc_dir)
                    
                # Удаляем оборудование из списка
                del self.equipment_data[found_index]

                with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.equipment_data, f, indent=4, ensure_ascii=False)

                self.update_equipment_table()
                self.update_dashboard()
                QMessageBox.information(self, "Успех", "Оборудование удалено")

            except Exception as e:
                QMessageBox.critical(
                    self, "Ошибка", f"Не удалось удалить оборудование: {str(e)}"
                )

    def export_equipment(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт данных",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)",
            options=options,
        )

        if not file_name:
            return

        try:
            if file_name.endswith(".xlsx"):
                self.export_to_excel(file_name)
            else:
                self.export_to_csv(file_name)

            QMessageBox.information(self, "Успех", "Данные успешно экспортированы")
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Не удалось экспортировать данные: {str(e)}"
            )

    def export_to_excel(self, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Оборудование"

        # Заголовки
        headers = [
            "Наименование",
            "Инвентарный номер",
            "Серийный номер",
            "Подразделение",
            "Дата поверки",
            "Следующая поверка",
            "Статус",
            "Примечания",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Данные
        today = datetime.now().date()

        for row, equipment in enumerate(self.equipment_data, 2):
            sheet.cell(row=row, column=1, value=equipment["name"])

            sheet.cell(row=row, column=2, value=equipment["inventory_number"])
            sheet.cell(row=row, column=3, value=equipment["serial_number"])
            sheet.cell(row=row, column=4, value=equipment["department"])

            verif_date = datetime.strptime(
                equipment["verification_date"], "%Y-%m-%d"
            ).date()
            sheet.cell(row=row, column=5, value=verif_date.strftime("%d.%m.%Y"))

            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            sheet.cell(row=row, column=6, value=next_verif_date.strftime("%d.%m.%Y"))

            days_left = (next_verif_date - today).days
            if days_left < 0:
                status = "Просрочено"
            elif days_left <= 30:
                status = f"Истекает через {days_left} дней"
            else:
                status = f"Активно ({days_left} дней)"

            sheet.cell(row=row, column=7, value=status)
            sheet.cell(row=row, column=8, value=equipment.get("notes", ""))

        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Границы для всех ячеек
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        workbook.save(file_path)

    def export_to_csv(self, file_path):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )

            # Заголовки
            writer.writerow(
                [
                    "Наименование",
                    "Инвентарный номер",
                    "Серийный номер",
                    "Подразделение",
                    "Дата поверки",
                    "Следующая поверка",
                    "Статус",
                    "Примечания",
                ]
            )

            # Данные
            today = datetime.now().date()

            for equipment in self.equipment_data:
                next_verif_date = datetime.strptime(
                    equipment["next_verification_date"], "%Y-%m-%d"
                ).date()
                days_left = (next_verif_date - today).days

                if days_left < 0:
                    status = "Просрочено"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                else:
                    status = f"Активно ({days_left} дней)"

                writer.writerow(
                    [
                        equipment["name"],
                        equipment["inventory_number"],
                        equipment["serial_number"],
                        equipment["department"],
                        datetime.strptime(
                            equipment["verification_date"], "%Y-%m-%d"
                        ).strftime("%d.%m.%Y"),
                        next_verif_date.strftime("%d.%m.%Y"),
                        status,
                        equipment.get("notes", ""),
                    ]
                )

    def generate_equipment_report(self, report_type):
        today = datetime.now().date()
        filtered_data = []

        department_filter = self.report_department_combo.currentText()
        if department_filter == "Все подразделения":
            department_filter = None

        for equipment in self.equipment_data:
            if department_filter and equipment["department"] != department_filter:
                continue

            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if report_type == "expiring" and not (0 <= days_left <= 30):
                continue
            elif report_type == "overdue" and days_left >= 0:
                continue

            filtered_data.append({"equipment": equipment, "days_left": days_left})

        # Сортируем данные
        if report_type in ["expiring", "overdue"]:
            filtered_data.sort(key=lambda x: x["days_left"])
        else:
            filtered_data.sort(key=lambda x: x["equipment"]["department"])

        # Обновляем предпросмотр
        self.equipment_report_preview.setRowCount(len(filtered_data))

        for row, data in enumerate(filtered_data):
            eq = data["equipment"]

            # Наименование
            name_item = QTableWidgetItem(eq["name"])
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.equipment_report_preview.setItem(row, 0, name_item)

            # Инвентарный номер
            inv_item = QTableWidgetItem(eq["inventory_number"])
            inv_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_report_preview.setItem(row, 1, inv_item)

            # Подразделение
            dept_item = QTableWidgetItem(eq["department"])
            dept_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_report_preview.setItem(row, 2, dept_item)

            # Дата поверки
            verif_date = datetime.strptime(eq["verification_date"], "%Y-%m-%d").date()
            verif_item = QTableWidgetItem(verif_date.strftime("%d.%m.%Y"))
            verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_report_preview.setItem(row, 3, verif_item)

            # Следующая поверка
            next_verif_date = datetime.strptime(
                eq["next_verification_date"], "%Y-%m-%d"
            ).date()
            next_verif_item = QTableWidgetItem(next_verif_date.strftime("%d.%m.%Y"))
            next_verif_item.setTextAlignment(Qt.AlignCenter)
            self.equipment_report_preview.setItem(row, 4, next_verif_item)

            # Статус
            if data["days_left"] < 0:
                status = "Просрочено"
                color = "#e74c3c"
            elif data["days_left"] <= 30:
                status = f"Истекает через {data['days_left']} дней"
                color = "#f39c12"
            else:
                status = f"Активно ({data['days_left']} дней)"
                color = "#2ecc71"

            status_item = QTableWidgetItem(status)
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setForeground(QColor(color))
            self.equipment_report_preview.setItem(row, 5, status_item)
            
        # Экспорт отчета
        if len(filtered_data) > 0:
            export_format = self.export_format_combo.currentData()
            report_title = ""
            
            if report_type == "expiring":
                report_title = "Истекающие поверки"
            elif report_type == "overdue":
                report_title = "Просроченные поверки"
            else:
                report_title = "Полный отчет по оборудованию"
                
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                f"Экспорт отчета: {report_title}",
                f"{report_title}",
                f"Excel Files (*.{export_format});;All Files (*)",
                options=options
            )
            
            if file_name:
                try:
                    if export_format == "xlsx":
                        self.export_equipment_report_to_excel(file_name, filtered_data, report_title)
                    elif export_format == "csv":
                        self.export_equipment_report_to_csv(file_name, filtered_data, report_title)
                    elif export_format == "pdf":
                        self.export_equipment_report_to_pdf(file_name, filtered_data, report_title)
                        
                    QMessageBox.information(self, "Успех", "Отчет успешно экспортирован")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта отчета: {str(e)}")
        else:
            QMessageBox.information(self, "Информация", "Нет данных для формирования отчета")
            
    def export_equipment_report_to_excel(self, file_path, data, title):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title
        
        # Заголовок отчета
        sheet.merge_cells('A1:F1')
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Дата отчета
        sheet.merge_cells('A2:F2')
        date_cell = sheet.cell(row=2, column=1, value=f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}")
        date_cell.alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = [
            "Наименование",
            "Инв. номер",
            "Подразделение",
            "Дата поверки",
            "След. поверка",
            "Статус",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            
        # Данные
        for row, item in enumerate(data, 5):
            eq = item["equipment"]
            days_left = item["days_left"]
            
            sheet.cell(row=row, column=1, value=eq["name"])
            sheet.cell(row=row, column=2, value=eq["inventory_number"])
            sheet.cell(row=row, column=3, value=eq["department"])
            
            verif_date = datetime.strptime(eq["verification_date"], "%Y-%m-%d").date()
            sheet.cell(row=row, column=4, value=verif_date.strftime("%d.%m.%Y"))
            
            next_verif_date = datetime.strptime(eq["next_verification_date"], "%Y-%m-%d").date()
            sheet.cell(row=row, column=5, value=next_verif_date.strftime("%d.%m.%Y"))
            
            if days_left < 0:
                status = "Просрочено"
            elif days_left <= 30:
                status = f"Истекает через {days_left} дней"
            else:
                status = f"Активно ({days_left} дней)"
                
            sheet.cell(row=row, column=6, value=status)
        
        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
            
        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                
        workbook.save(file_path)
            
    def export_equipment_report_to_csv(self, file_path, data, title):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            
            # Заголовок отчета
            writer.writerow([title])
            writer.writerow([f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}"])
            writer.writerow([])  # Пустая строка
            
            # Заголовки таблицы
            writer.writerow([
                "Наименование",
                "Инв. номер",
                "Подразделение",
                "Дата поверки",
                "След. поверка",
                "Статус",
            ])
            
            # Данные
            for item in data:
                eq = item["equipment"]
                days_left = item["days_left"]
                
                if days_left < 0:
                    status = "Просрочено"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                else:
                    status = f"Активно ({days_left} дней)"
                    
                writer.writerow([
                    eq["name"],
                    eq["inventory_number"],
                    eq["department"],
                    datetime.strptime(eq["verification_date"], "%Y-%m-%d").strftime("%d.%m.%Y"),
                    datetime.strptime(eq["next_verification_date"], "%Y-%m-%d").strftime("%d.%m.%Y"),
                    status,
                ])
                
    def export_equipment_report_to_pdf(self, file_path, data, title):
        # Создаем простой HTML для преобразования в PDF
        html = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ text-align: center; }}
                .date {{ text-align: center; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; text-align: center; }}
                .expired {{ color: #e74c3c; }}
                .expiring {{ color: #f39c12; }}
                .active {{ color: #2ecc71; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <div class="date">Дата формирования: {datetime.now().strftime('%d.%m.%Y')}</div>
            
            <table>
                <tr>
                    <th>Наименование</th>
                    <th>Инв. номер</th>
                    <th>Подразделение</th>
                    <th>Дата поверки</th>
                    <th>След. поверка</th>
                    <th>Статус</th>
                </tr>
        """
        
        for item in data:
            eq = item["equipment"]
            days_left = item["days_left"]
            
            if days_left < 0:
                status = "Просрочено"
                status_class = "expired"
            elif days_left <= 30:
                status = f"Истекает через {days_left} дней"
                status_class = "expiring"
            else:
                status = f"Активно ({days_left} дней)"
                status_class = "active"
                
            html += f"""
                <tr>
                    <td>{eq["name"]}</td>
                    <td>{eq["inventory_number"]}</td>
                    <td>{eq["department"]}</td>
                    <td>{datetime.strptime(eq["verification_date"], "%Y-%m-%d").strftime("%d.%m.%Y")}</td>
                    <td>{datetime.strptime(eq["next_verification_date"], "%Y-%m-%d").strftime("%d.%m.%Y")}</td>
                    <td class="{status_class}">{status}</td>
                </tr>
            """
            
        html += """
            </table>
        </body>
        </html>
        """
        
        # Используем QPrinter для создания PDF
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(file_path)
        printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
        
        # Создаем документ и печатаем в PDF
        document = QTextEdit()
        document.setHtml(html)
        document.print_(printer)

    def generate_specialists_report(self, report_type):
        today = datetime.now().date()
        filtered_data = []

        department_filter = self.spec_report_department_combo.currentText()
        method_filter = self.spec_report_method_combo.currentText()
        
        if department_filter == "Все подразделения":
            department_filter = None
            
        if method_filter == "Все методы":
            method_filter = None
            
        for specialist in self.specialists_data:
            # Применяем фильтр по подразделению
            if department_filter and specialist.get("department") != department_filter:
                continue
                
            # Применяем фильтр по методу контроля
            if method_filter and method_filter not in specialist.get("methods", {}):
                continue
                
            # Для отчета по истекающим удостоверениям
            if report_type == "expiring":
                if "next_cert_date" not in specialist:
                    continue
                    
                next_cert_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                days_left = (next_cert_date - today).days
                
                if not (days_left >= 0 and days_left <= 30):
                    continue
                    
            filtered_data.append(specialist)
            
        # Сортируем данные
        if report_type == "expiring":
            # Сортировка по дате истечения (ближайшие сначала)
            filtered_data.sort(key=lambda x: datetime.strptime(x["next_cert_date"], "%Y-%m-%d"))
        elif report_type == "methods":
            # Сортировка по методам контроля
            filtered_data.sort(key=lambda x: len(x.get("methods", {})), reverse=True)
        else:
            # Сортировка по фамилии
            filtered_data.sort(key=lambda x: x.get("last_name", ""))
            
        # Обновляем предпросмотр
        self.specialists_report_preview.setRowCount(len(filtered_data))
        
        for row, specialist in enumerate(filtered_data):
            # ФИО
            name_item = QTableWidgetItem(specialist.get("full_name", ""))
            self.specialists_report_preview.setItem(row, 0, name_item)
            
            # Должность
            position_item = QTableWidgetItem(specialist.get("position", ""))
            self.specialists_report_preview.setItem(row, 1, position_item)
            
            # Подразделение
            dept_item = QTableWidgetItem(specialist.get("department", ""))
            self.specialists_report_preview.setItem(row, 2, dept_item)
            
            # Методы контроля
            methods = specialist.get("methods", {})
            methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
            methods_item = QTableWidgetItem(methods_text)
            self.specialists_report_preview.setItem(row, 3, methods_item)
            
            # Срок действия
            if "next_cert_date" in specialist:
                next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                next_date_item = QTableWidgetItem(next_date.strftime("%d.%m.%Y"))
                self.specialists_report_preview.setItem(row, 4, next_date_item)
                
                # Статус
                days_left = (next_date - today).days
                
                if days_left < 0:
                    status = "Просрочено"
                    color = "#e74c3c"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                    color = "#f39c12"
                else:
                    status = f"Активно ({days_left} дней)"
                    color = "#2ecc71"
                    
                status_item = QTableWidgetItem(status)
                status_item.setForeground(QColor(color))
                self.specialists_report_preview.setItem(row, 5, status_item)
            else:
                self.specialists_report_preview.setItem(row, 4, QTableWidgetItem("Не указан"))
                self.specialists_report_preview.setItem(row, 5, QTableWidgetItem("Не указан"))
                
        # Экспорт отчета
        if len(filtered_data) > 0:
            export_format = self.spec_export_format_combo.currentData()
            report_title = ""
            
            if report_type == "expiring":
                report_title = "Истекающие удостоверения специалистов"
            elif report_type == "methods":
                report_title = "Отчет по методам контроля специалистов"
            else:
                report_title = "Полный отчет по специалистам"
                
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                f"Экспорт отчета: {report_title}",
                f"{report_title}",
                f"Excel Files (*.{export_format});;All Files (*)",
                options=options
            )
            
            if file_name:
                try:
                    if export_format == "xlsx":
                        self.export_specialists_report_to_excel(file_name, filtered_data, report_title)
                    elif export_format == "csv":
                        self.export_specialists_report_to_csv(file_name, filtered_data, report_title)
                    elif export_format == "pdf":
                        self.export_specialists_report_to_pdf(file_name, filtered_data, report_title)
                        
                    QMessageBox.information(self, "Успех", "Отчет успешно экспортирован")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта отчета: {str(e)}")
        else:
            QMessageBox.information(self, "Информация", "Нет данных для формирования отчета")
            
    def export_specialists_report_to_excel(self, file_path, data, title):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Специалисты"
        
        # Заголовок отчета
        sheet.merge_cells('A1:F1')
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Дата отчета
        sheet.merge_cells('A2:F2')
        date_cell = sheet.cell(row=2, column=1, value=f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}")
        date_cell.alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = [
            "ФИО",
            "Должность",
            "Подразделение",
            "Методы контроля",
            "Срок действия",
            "Статус",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            
        # Данные
        today = datetime.now().date()
        
        for row, specialist in enumerate(data, 5):
            sheet.cell(row=row, column=1, value=specialist.get("full_name", ""))
            sheet.cell(row=row, column=2, value=specialist.get("position", ""))
            sheet.cell(row=row, column=3, value=specialist.get("department", ""))
            
            methods = specialist.get("methods", {})
            methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
            sheet.cell(row=row, column=4, value=methods_text)
            
            if "next_cert_date" in specialist:
                next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                sheet.cell(row=row, column=5, value=next_date.strftime("%d.%m.%Y"))
                
                days_left = (next_date - today).days
                
                if days_left < 0:
                    status = "Просрочено"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                else:
                    status = f"Активно ({days_left} дней)"
                    
                sheet.cell(row=row, column=6, value=status)
            else:
                sheet.cell(row=row, column=5, value="Не указан")
                sheet.cell(row=row, column=6, value="Не указан")
        
        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
            
        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                
        workbook.save(file_path)
            
    def export_specialists_report_to_csv(self, file_path, data, title):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            
            # Заголовок отчета
            writer.writerow([title])
            writer.writerow([f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}"])
            writer.writerow([])  # Пустая строка
            
            # Заголовки таблицы
            writer.writerow([
                "ФИО",
                "Должность",
                "Подразделение",
                "Методы контроля",
                "Срок действия",
                "Статус",
            ])
            
            # Данные
            today = datetime.now().date()
            
            for specialist in data:
                methods = specialist.get("methods", {})
                methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
                
                status = ""
                next_date_str = ""
                
                if "next_cert_date" in specialist:
                    next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                    next_date_str = next_date.strftime("%d.%m.%Y")
                    
                    days_left = (next_date - today).days
                    
                    if days_left < 0:
                        status = "Просрочено"
                    elif days_left <= 30:
                        status = f"Истекает через {days_left} дней"
                    else:
                        status = f"Активно ({days_left} дней)"
                else:
                    next_date_str = "Не указан"
                    status = "Не указан"
                    
                writer.writerow([
                    specialist.get("full_name", ""),
                    specialist.get("position", ""),
                    specialist.get("department", ""),
                    methods_text,
                    next_date_str,
                    status,
                ])
                
    def export_specialists_report_to_pdf(self, file_path, data, title):
        # Создаем простой HTML для преобразования в PDF
        html = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ text-align: center; }}
                .date {{ text-align: center; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; text-align: center; }}
                .expired {{ color: #e74c3c; }}
                .expiring {{ color: #f39c12; }}
                .active {{ color: #2ecc71; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <div class="date">Дата формирования: {datetime.now().strftime('%d.%m.%Y')}</div>
            
            <table>
                <tr>
                    <th>ФИО</th>
                    <th>Должность</th>
                    <th>Подразделение</th>
                    <th>Методы контроля</th>
                    <th>Срок действия</th>
                    <th>Статус</th>
                </tr>
        """
        
        today = datetime.now().date()
        
        for specialist in data:
            methods = specialist.get("methods", {})
            methods_text = ", ".join([f"{m} ({lvl})" for m, lvl in methods.items()])
            
            status = ""
            status_class = ""
            next_date_str = ""
            
            if "next_cert_date" in specialist:
                next_date = datetime.strptime(specialist["next_cert_date"], "%Y-%m-%d").date()
                next_date_str = next_date.strftime("%d.%m.%Y")
                
                days_left = (next_date - today).days
                
                if days_left < 0:
                    status = "Просрочено"
                    status_class = "expired"
                elif days_left <= 30:
                    status = f"Истекает через {days_left} дней"
                    status_class = "expiring"
                else:
                    status = f"Активно ({days_left} дней)"
                    status_class = "active"
            else:
                next_date_str = "Не указан"
                status = "Не указан"
                
            html += f"""
                <tr>
                    <td>{specialist.get("full_name", "")}</td>
                    <td>{specialist.get("position", "")}</td>
                    <td>{specialist.get("department", "")}</td>
                    <td>{methods_text}</td>
                    <td>{next_date_str}</td>
                    <td class="{status_class}">{status}</td>
                </tr>
            """
            
        html += """
            </table>
        </body>
        </html>
        """
        
        # Используем QPrinter для создания PDF
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(file_path)
        printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
        
        # Создаем документ и печатаем в PDF
        document = QTextEdit()
        document.setHtml(html)
        document.print_(printer)

    def generate_objects_report(self, report_type):
        filtered_data = []

        department_filter = self.obj_report_department_combo.currentText()
        result_filter = self.obj_report_result_combo.currentText()
        start_date = self.obj_report_start_date.date().toPyDate()
        end_date = self.obj_report_end_date.date().toPyDate()
        
        if department_filter == "Все подразделения":
            department_filter = None
            
        if result_filter == "Все результаты":
            result_filter = None
            
        for obj in self.objects_data:
            # Применяем фильтр по подразделению
            if department_filter and obj.get("department") != department_filter:
                continue
                
            # Применяем фильтр по результату
            if result_filter:
                if result_filter == "Нет заключения":
                    if "conclusion" in obj and obj["conclusion"].get("result"):
                        continue
                elif "conclusion" not in obj or obj["conclusion"].get("result") != result_filter:
                    continue
                    
            # Применяем фильтр по дате
            if "date" in obj:
                obj_date = datetime.strptime(obj["date"], "%Y-%m-%d").date()
                if obj_date < start_date or obj_date > end_date:
                    continue
                    
            filtered_data.append(obj)
            
        # Сортируем данные
        if report_type == "results":
            # Сортировка по результату
            filtered_data.sort(key=lambda x: x.get("conclusion", {}).get("result", ""))
        elif report_type == "customers":
            # Сортировка по заказчику
            filtered_data.sort(key=lambda x: x.get("customer", ""))
        else:
            # Сортировка по дате (от новых к старым)
            filtered_data.sort(key=lambda x: x.get("date", ""), reverse=True)
            
        # Обновляем предпросмотр
        self.objects_report_preview.setRowCount(len(filtered_data))
        
        for row, obj in enumerate(filtered_data):
            # Наименование
            name_item = QTableWidgetItem(obj.get("name", ""))
            self.objects_report_preview.setItem(row, 0, name_item)
            
            # Номер
            number_item = QTableWidgetItem(obj.get("number", ""))
            self.objects_report_preview.setItem(row, 1, number_item)
            
            # Дата
            if "date" in obj:
                date_obj = datetime.strptime(obj["date"], "%Y-%m-%d").date()
                date_item = QTableWidgetItem(date_obj.strftime("%d.%m.%Y"))
                self.objects_report_preview.setItem(row, 2, date_item)
            else:
                self.objects_report_preview.setItem(row, 2, QTableWidgetItem(""))
                
            # Заказчик
            customer_item = QTableWidgetItem(obj.get("customer", ""))
            self.objects_report_preview.setItem(row, 3, customer_item)
            
            # Номер заключения и результат
            conclusion = obj.get("conclusion", {})
            if conclusion:
                conclusion_number = QTableWidgetItem(conclusion.get("number", ""))
                self.objects_report_preview.setItem(row, 4, conclusion_number)
                
                result_item = QTableWidgetItem(conclusion.get("result", ""))
                # Цветовое оформление результата
                if conclusion.get("result") == "Годен":
                    result_item.setForeground(QColor("#2ecc71"))
                elif conclusion.get("result") == "Не годен":
                    result_item.setForeground(QColor("#e74c3c"))
                elif conclusion.get("result") == "Требуется доработка":
                    result_item.setForeground(QColor("#f39c12"))
                    
                self.objects_report_preview.setItem(row, 5, result_item)
            else:
                self.objects_report_preview.setItem(row, 4, QTableWidgetItem(""))
                no_conclusion = QTableWidgetItem("Нет заключения")
                no_conclusion.setForeground(QColor("#7f8c8d"))
                self.objects_report_preview.setItem(row, 5, no_conclusion)
                
        # Экспорт отчета
        if len(filtered_data) > 0:
            export_format = self.obj_export_format_combo.currentData()
            report_title = ""
            
            if report_type == "results":
                report_title = "Отчет по результатам контроля объектов"
            elif report_type == "customers":
                report_title = "Отчет по заказчикам"
            else:
                report_title = "Полный отчет по объектам"
                
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                f"Экспорт отчета: {report_title}",
                f"{report_title}",
                f"Excel Files (*.{export_format});;All Files (*)",
                options=options
            )
            
            if file_name:
                try:
                    if export_format == "xlsx":
                        self.export_objects_report_to_excel(file_name, filtered_data, report_title)
                    elif export_format == "csv":
                        self.export_objects_report_to_csv(file_name, filtered_data, report_title)
                    elif export_format == "pdf":
                        self.export_objects_report_to_pdf(file_name, filtered_data, report_title)
                        
                    QMessageBox.information(self, "Успех", "Отчет успешно экспортирован")
                except Exception as e:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта отчета: {str(e)}")
        else:
            QMessageBox.information(self, "Информация", "Нет данных для формирования отчета")
            
    def export_objects_report_to_excel(self, file_path, data, title):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Объекты"
        
        # Заголовок отчета
        sheet.merge_cells('A1:F1')
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Дата отчета
        sheet.merge_cells('A2:F2')
        date_cell = sheet.cell(row=2, column=1, value=f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}")
        date_cell.alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = [
            "Наименование",
            "Номер",
            "Дата",
            "Заказчик",
            "Номер заключения",
            "Результат",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            
        # Данные
        for row, obj in enumerate(data, 5):
            sheet.cell(row=row, column=1, value=obj.get("name", ""))
            sheet.cell(row=row, column=2, value=obj.get("number", ""))
            
            if "date" in obj:
                date_obj = datetime.strptime(obj["date"], "%Y-%m-%d").date()
                sheet.cell(row=row, column=3, value=date_obj.strftime("%d.%m.%Y"))
                
            sheet.cell(row=row, column=4, value=obj.get("customer", ""))
            
            conclusion = obj.get("conclusion", {})
            if conclusion:
                sheet.cell(row=row, column=5, value=conclusion.get("number", ""))
                sheet.cell(row=row, column=6, value=conclusion.get("result", ""))
            else:
                sheet.cell(row=row, column=6, value="Нет заключения")
        
        # Форматирование
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
            
        # Границы
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                
        workbook.save(file_path)
            
    def export_objects_report_to_csv(self, file_path, data, title):
        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL
            )
            
            # Заголовок отчета
            writer.writerow([title])
            writer.writerow([f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}"])
            writer.writerow([])  # Пустая строка
            
            # Заголовки таблицы
            writer.writerow([
                "Наименование",
                "Номер",
                "Дата",
                "Заказчик",
                "Номер заключения",
                "Результат",
            ])
            
            # Данные
            for obj in data:
                date_str = ""
                if "date" in obj:
                    date_str = datetime.strptime(obj["date"], "%Y-%m-%d").strftime("%d.%m.%Y")
                    
                conclusion = obj.get("conclusion", {})
                conclusion_number = conclusion.get("number", "") if conclusion else ""
                result = conclusion.get("result", "") if conclusion else "Нет заключения"
                    
                writer.writerow([
                    obj.get("name", ""),
                    obj.get("number", ""),
                    date_str,
                    obj.get("customer", ""),
                    conclusion_number,
                    result,
                ])
                
    def export_objects_report_to_pdf(self, file_path, data, title):
        # Создаем простой HTML для преобразования в PDF
        html = f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ text-align: center; }}
                .date {{ text-align: center; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; text-align: center; }}
                .success {{ color: #2ecc71; }}
                .danger {{ color: #e74c3c; }}
                .warning {{ color: #f39c12; }}
                .neutral {{ color: #7f8c8d; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <div class="date">Дата формирования: {datetime.now().strftime('%d.%m.%Y')}</div>
            
            <table>
                <tr>
                    <th>Наименование</th>
                    <th>Номер</th>
                    <th>Дата</th>
                    <th>Заказчик</th>
                    <th>Номер заключения</th>
                    <th>Результат</th>
                </tr>
        """
        
        for obj in data:
            date_str = ""
            if "date" in obj:
                date_str = datetime.strptime(obj["date"], "%Y-%m-%d").strftime("%d.%m.%Y")
                
            conclusion = obj.get("conclusion", {})
            conclusion_number = conclusion.get("number", "") if conclusion else ""
            
            result = ""
            result_class = ""
            
            if conclusion and "result" in conclusion:
                result = conclusion["result"]
                if result == "Годен":
                    result_class = "success"
                elif result == "Не годен":
                    result_class = "danger"
                elif result == "Требуется доработка":
                    result_class = "warning"
            else:
                result = "Нет заключения"
                result_class = "neutral"
                
            html += f"""
                <tr>
                    <td>{obj.get("name", "")}</td>
                    <td>{obj.get("number", "")}</td>
                    <td>{date_str}</td>
                    <td>{obj.get("customer", "")}</td>
                    <td>{conclusion_number}</td>
                    <td class="{result_class}">{result}</td>
                </tr>
            """
            
        html += """
            </table>
        </body>
        </html>
        """
        
        # Используем QPrinter для создания PDF
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(file_path)
        printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
        
        # Создаем документ и печатаем в PDF
        document = QTextEdit()
        document.setHtml(html)
        document.print_(printer)

    def check_verification_dates(self):
        today = datetime.now().date()
        expiring_soon = []
        overdue = []

        for equipment in self.equipment_data:
            next_verif_date = datetime.strptime(
                equipment["next_verification_date"], "%Y-%m-%d"
            ).date()
            days_left = (next_verif_date - today).days

            if days_left < 0:
                overdue.append(equipment)
            elif days_left <= 7:  # Уведомляем за неделю до поверки
                expiring_soon.append(equipment)
                
        # Проверяем удостоверения специалистов
        expiring_certs = []
        overdue_certs = []
        
        for specialist in self.specialists_data:
            if "next_cert_date" in specialist:
                next_cert_date = datetime.strptime(
                    specialist["next_cert_date"], "%Y-%m-%d"
                ).date()
                days_left = (next_cert_date - today).days
                
                if days_left < 0:
                    overdue_certs.append(specialist)
                elif days_left <= 7:  # Уведомляем за неделю
                    expiring_certs.append(specialist)

        if overdue or expiring_soon or overdue_certs or expiring_certs:
            message = ""

            if overdue:
                message += "<b>Просроченные поверки:</b><br>"
                for eq in overdue:
                    message += f"- {eq['name']} (инв. № {eq['inventory_number']}, {eq['department']})<br>"
                message += "<br>"

            if expiring_soon:
                message += "<b>Скоро истекают поверки:</b><br>"
                for eq in expiring_soon:
                    next_date = datetime.strptime(
                        eq["next_verification_date"], "%Y-%m-%d"
                    ).date()
                    days_left = (next_date - today).days
                    message += f"- {eq['name']} (инв. № {eq['inventory_number']}, {eq['department']}) - осталось {days_left} дней<br>"
                message += "<br>"
                
            if overdue_certs:
                message += "<b>Просроченные удостоверения:</b><br>"
                for spec in overdue_certs:
                    message += f"- {spec.get('full_name', '')} ({spec.get('position', '')}, {spec.get('department', '')})<br>"
                message += "<br>"
                
            if expiring_certs:
                message += "<b>Скоро истекают удостоверения:</b><br>"
                for spec in expiring_certs:
                    next_date = datetime.strptime(
                        spec["next_cert_date"], "%Y-%m-%d"
                    ).date()
                    days_left = (next_date - today).days
                    message += f"- {spec.get('full_name', '')} ({spec.get('position', '')}, {spec.get('department', '')}) - осталось {days_left} дней<br>"

            # Показываем уведомление только если пользователь активен
            if self.isActiveWindow():
                QMessageBox.warning(self, "Проверка сроков", message, QMessageBox.Ok)

            # Отправляем уведомление на email, если он указан
            if self.user_email and (overdue or expiring_soon or overdue_certs or expiring_certs):
                self.send_email_notification(overdue, expiring_soon, overdue_certs, expiring_certs)

    def send_email_notification(self, overdue, expiring_soon, overdue_certs, expiring_certs):
        # В реальном приложении здесь должна быть реализация отправки email
        # Это примерная заглушка для демонстрации
        print(f"Уведомление отправлено на email: {self.user_email}")
        print("Просроченные поверки:", overdue)
        print("Скоро истекают:", expiring_soon)
        print("Просроченные удостоверения:", overdue_certs)
        print("Скоро истекают удостоверения:", expiring_certs)

    def show_dashboard(self):
        self.content_area.setCurrentWidget(self.dashboard_page)
        self.update_dashboard()

    def show_equipment(self):
        self.content_area.setCurrentWidget(self.equipment_page)

    def show_specialists(self):
        self.content_area.setCurrentWidget(self.specialists_page)
        
    def show_objects(self):
        self.content_area.setCurrentWidget(self.objects_page)

    def show_reports(self):
        self.content_area.setCurrentWidget(self.reports_page)

    def show_ai_chat(self):
        self.content_area.setCurrentWidget(self.ai_chat_page)

    def manage_users(self):
        if self.user_role != "admin":
            QMessageBox.warning(
                self, "Ошибка", "Недостаточно прав для управления пользователями"
            )
            return

        dialog = UserManagementDialog(self)
        dialog.exec_()

    def close_application(self):
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите выйти?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply == QMessageBox.Yes:
            QCoreApplication.quit()

def initialize_data_files():
    # Создаем файл пользователей, если его нет
    if not os.path.exists(USER_DATA_FILE):
        default_users = {
            "admin": {
                "password": sha256("admin123".encode("utf-8")).hexdigest(),
                "role": "admin",
                "email": "admin@example.com",
            }
        }

        with open(USER_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(default_users, f, indent=4, ensure_ascii=False)

    # Создаем файл оборудования, если его нет
    if not os.path.exists(EQUIPMENT_DATA_FILE):
        with open(EQUIPMENT_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
            
    # Создаем файл специалистов, если его нет
    if not os.path.exists(SPECIALISTS_DATA_FILE):
        with open(SPECIALISTS_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
            
    # Создаем файл объектов, если его нет
    if not os.path.exists(OBJECTS_DATA_FILE):
        with open(OBJECTS_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
            
    # Создаем файл конфигурации, если его нет
    if not os.path.exists(CONFIG_FILE):
        config = {
            "departments": DEPARTMENTS,
            "control_methods": CONTROL_METHODS,
            "equipment_types": [
                "Манометр", "Ультразвуковой дефектоскоп", "Толщиномер", 
                "Течеискатель", "Рентгеновский аппарат", "Вихретоковый дефектоскоп"
            ]
        }
        
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
            
    # Создаем директорию для документов, если её нет
    if not os.path.exists(DOCS_FOLDER):
        os.makedirs(DOCS_FOLDER)

def main():
    # Инициализация файлов данных
    initialize_data_files()

    # Создание приложения
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Авторизация пользователя
    login_dialog = LoginDialog()
    if login_dialog.exec_() == QDialog.Accepted:
        # Запуск основного приложения
        main_app = EquipmentControlApp(
            user_role=login_dialog.user_role,
            user_email=login_dialog.user_email,
            username=login_dialog.username,
        )
        main_app.show()
        sys.exit(app.exec_())
    else:
        sys.exit(0)

if __name__ == "__main__":
    main()